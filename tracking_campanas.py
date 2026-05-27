"""
tracking_campanas.py
Análisis de cohortes y tracking de damas a lo largo de 10 campañas.
Genera un Excel ejecutivo con 4 pestañas.
Incluye estado "Inactiva" (IdSituacion=0 / INICIAL) con tracking de transición.
"""

import warnings
warnings.filterwarnings("ignore")

from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ── Configuración ─────────────────────────────────────────────────────────────
INPUT_FILE = Path("/root/.claude/uploads/48129c5c-b81b-412d-8c25-6859dcfca1e5/ba7a1adc-GENERAL_CARTERA_MORAS_CAMPA_A_1__10_2.xlsx")
OUTPUT_FILE = Path("/home/user/Arabela/reporte_cohortes_campanas.xlsx")
SHEET_NAME  = "Consolidado"
CHURN_WINDOW = 3   # campañas sin aparecer → fuga

# Paleta ejecutiva
COLOR_HEADER   = "1A3C6E"   # azul oscuro
COLOR_SUBHEAD  = "2563EB"   # azul medio
COLOR_INACTIVA = "E2E8F0"   # gris azulado pálido
COLOR_MORA1    = "FEF9C3"   # amarillo pálido
COLOR_MORA2    = "FFEDD5"   # naranja pálido
COLOR_MORA3    = "FEE2E2"   # rojo pálido
COLOR_ALT      = "F1F5F9"   # gris claro (filas alternas)
COLOR_TOTAL    = "DBEAFE"   # azul claro (totales)
COLOR_WHITE    = "FFFFFF"
MORA_LEVELS    = ["Inactiva", "Mora 1", "Mora 2", "Mora 3"]

# ── Estilos helpers ───────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border_thin():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _write(ws, row, col, value, bold=False, fill_hex=None,
           font_color="000000", font_size=10, align_h="left",
           num_format=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, color=font_color, size=font_size)
    cell.alignment = _align(h=align_h, wrap=wrap)
    cell.border    = _border_thin()
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if num_format:
        cell.number_format = num_format
    return cell

def _header(ws, row, col, value, colspan=1, color=COLOR_HEADER):
    cell = _write(ws, row, col, value, bold=True,
                  fill_hex=color, font_color=COLOR_WHITE,
                  font_size=11, align_h="center")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
    return cell

def _pct(a, b):
    return round(a / b * 100, 1) if b else 0.0

# ── Carga y preparación ───────────────────────────────────────────────────────
def load_data() -> pd.DataFrame:
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
    df.columns = [c.strip() for c in df.columns]
    df["NoDama"]    = df["NoDama"].astype(str).str.strip()
    df["Campania"]  = df["Campania"].astype(str).str.strip().str.upper()
    df["Moras"]     = df["Moras"].astype(str).str.strip()
    df["SaldoDama"] = pd.to_numeric(df["SaldoDama"], errors="coerce").fillna(0)
    df["_camp_n"]   = df["Campania"].str.replace("C", "", regex=False).apply(
                          lambda x: int(x) if x.isdigit() else None)
    df = df.dropna(subset=["_camp_n"]).copy()
    df["_camp_n"] = df["_camp_n"].astype(int)

    # Cuentas con IdSituacion=0 (INICIAL) → estado "Inactiva"
    if "IdSituacion" in df.columns:
        mask = df["IdSituacion"].astype(str).str.strip() == "0"
        df.loc[mask, "Moras"] = "Inactiva"
        n_inactivas = mask.sum()
        print(f"   ℹ️  {n_inactivas:,} registros reclasificados como 'Inactiva' (IdSituacion=0)")

    return df

# ── Motor de análisis ─────────────────────────────────────────────────────────
def build_analysis(df: pd.DataFrame) -> dict:
    camps = sorted(df["_camp_n"].unique())          # [1..10]
    labels = {c: f"C{c}" for c in camps}

    # Conjunto de NoDamas por campaña
    sets: dict[int, set] = {c: set(df[df["_camp_n"] == c]["NoDama"]) for c in camps}

    # Primera campaña en que aparece cada dama
    first_camp: dict[str, int] = (
        df.groupby("NoDama")["_camp_n"].min().to_dict()
    )

    # Datos por (NoDama, _camp_n) → mora y saldo
    rec = (
        df.sort_values("_camp_n", ascending=False)
          .drop_duplicates(subset=["NoDama", "_camp_n"])
          .set_index(["NoDama", "_camp_n"])
    )

    # Sets por mora y campaña para lookups rápidos
    mora_sets: dict[tuple, set] = {}
    for c in camps:
        df_c = df[df["_camp_n"] == c]
        for mora in MORA_LEVELS:
            mora_sets[(c, mora)] = set(df_c[df_c["Moras"] == mora]["NoDama"])

    results = []

    for i, c in enumerate(camps):
        prev_c = camps[i - 1] if i > 0 else None
        churn_target = c + CHURN_WINDOW

        total_set = sets[c]
        nuevas    = {d for d in total_set if first_camp.get(d) == c}
        retenidas = (total_set & sets[prev_c]) if prev_c else set()

        # Fuga: damas en C que no aparecen en ninguna de las próximas CHURN_WINDOW campañas
        future_camps = [cx for cx in camps if cx > c and cx <= churn_target]
        if future_camps:
            appeared_future = set().union(*[sets[cx] for cx in future_camps])
            fugadas = total_set - appeared_future
        else:
            fugadas = set()

        row_base = {
            "camp_n":        c,
            "camp_label":    labels[c],
            "total":         len(total_set),
            "nuevas":        len(nuevas),
            "retenidas":     len(retenidas),
            "fugadas":       len(fugadas),
            "saldo_total":   df[df["_camp_n"] == c]["SaldoDama"].sum(),
        }

        # Desglose por mora
        df_c = df[df["_camp_n"] == c]
        for mora in MORA_LEVELS:
            ids_m = mora_sets[(c, mora)]
            nuevas_m    = ids_m & nuevas
            retenidas_m = ids_m & retenidas
            fugadas_m   = ids_m & fugadas
            row_base[f"{mora}_total"]     = len(ids_m)
            row_base[f"{mora}_nuevas"]    = len(nuevas_m)
            row_base[f"{mora}_retenidas"] = len(retenidas_m)
            row_base[f"{mora}_fugadas"]   = len(fugadas_m)
            row_base[f"{mora}_saldo"]     = df_c[df_c["Moras"] == mora]["SaldoDama"].sum()

        # Transición desde Inactiva del período anterior → mora actual
        if prev_c is not None:
            inactivas_prev = mora_sets[(prev_c, "Inactiva")]
            for mora in MORA_LEVELS:
                # cuántas que eran Inactiva en prev_c ahora tienen este estado en c
                row_base[f"from_inactiva_to_{mora}"] = len(inactivas_prev & mora_sets[(c, mora)])
            row_base["from_inactiva_total"] = len(
                inactivas_prev & total_set
            )
        else:
            for mora in MORA_LEVELS:
                row_base[f"from_inactiva_to_{mora}"] = 0
            row_base["from_inactiva_total"] = 0

        results.append(row_base)

    summary_df = pd.DataFrame(results)

    # Matriz de transición de mora (camp n → camp n+1, para damas en ambas)
    transitions = []
    for i in range(len(camps) - 1):
        c1, c2 = camps[i], camps[i + 1]
        shared = sets[c1] & sets[c2]
        df1 = df[(df["_camp_n"] == c1) & df["NoDama"].isin(shared)].set_index("NoDama")["Moras"]
        df2 = df[(df["_camp_n"] == c2) & df["NoDama"].isin(shared)].set_index("NoDama")["Moras"]
        joined = df1.to_frame("mora_origen").join(df2.to_frame("mora_destino"), how="inner")
        for (orig, dest), cnt in joined.groupby(["mora_origen", "mora_destino"]).size().items():
            transitions.append({
                "de":     labels[c1], "a": labels[c2],
                "origen": orig, "destino": dest, "cuentas": cnt,
            })
    trans_df = pd.DataFrame(transitions) if transitions else pd.DataFrame()

    # Damas que salen definitivamente (no vuelven tras su última campaña)
    last_camp: dict[str, int] = df.groupby("NoDama")["_camp_n"].max().to_dict()
    max_c = max(camps)
    exits = []
    for dama, lc in last_camp.items():
        if lc < max_c:
            mora_exit = "N/A"
            try:
                mora_exit = rec.loc[(dama, lc), "Moras"]
            except Exception:
                pass
            saldo_exit = 0
            try:
                saldo_exit = float(rec.loc[(dama, lc), "SaldoDama"])
            except Exception:
                pass
            exits.append({"NoDama": dama, "UltimaCampaña": f"C{lc}",
                          "MoraAlSalir": mora_exit, "SaldoAlSalir": saldo_exit})
    exits_df = pd.DataFrame(exits)

    return {
        "summary":     summary_df,
        "transitions": trans_df,
        "exits":       exits_df,
        "camps":       camps,
        "labels":      labels,
    }

# ── Pestaña 1: Resumen Ejecutivo ──────────────────────────────────────────────
def write_resumen(wb: Workbook, data: dict):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 16

    summary = data["summary"]
    camps   = data["camps"]

    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value     = "ANÁLISIS DE COHORTES — COMPORTAMIENTO DE CARTERA POR CAMPAÑA"
    cell.font      = Font(bold=True, size=14, color=COLOR_WHITE)
    cell.fill      = _fill(COLOR_HEADER)
    cell.alignment = _align(h="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"].value     = f"Campañas C1–C{max(camps)} · Ventana de fuga: {CHURN_WINDOW} campañas · Estado 'Inactiva' = IdSituacion 0 (INICIAL)"
    ws["A2"].font      = Font(italic=True, size=10, color="64748B")
    ws["A2"].alignment = _align(h="center")

    row = 4

    mora_colors_hex = {
        "Inactiva": COLOR_INACTIVA,
        "Mora 1":   COLOR_MORA1,
        "Mora 2":   COLOR_MORA2,
        "Mora 3":   COLOR_MORA3,
    }

    for _, r in summary.iterrows():
        lbl = r["camp_label"]
        tot = int(r["total"])

        # Encabezado campaña
        ws.merge_cells(f"A{row}:J{row}")
        cell = ws[f"A{row}"]
        cell.value     = f"  CAMPAÑA {lbl}"
        cell.font      = Font(bold=True, size=12, color=COLOR_WHITE)
        cell.fill      = _fill(COLOR_SUBHEAD)
        cell.alignment = _align(h="left")
        ws.row_dimensions[row].height = 22
        row += 1

        # KPIs generales
        _write(ws, row, 1, "Total damas", bold=True, fill_hex=COLOR_ALT)
        _write(ws, row, 2, tot, fill_hex=COLOR_WHITE, align_h="center")
        _write(ws, row, 3, "Nuevas", bold=True, fill_hex=COLOR_ALT)
        _write(ws, row, 4, int(r["nuevas"]), fill_hex=COLOR_WHITE, align_h="center")
        _write(ws, row, 5, f"{_pct(r['nuevas'], tot)}%", fill_hex=COLOR_WHITE, align_h="center")
        _write(ws, row, 6, "Retenidas", bold=True, fill_hex=COLOR_ALT)
        _write(ws, row, 7, int(r["retenidas"]), fill_hex=COLOR_WHITE, align_h="center")
        _write(ws, row, 8, f"{_pct(r['retenidas'], tot)}%", fill_hex=COLOR_WHITE, align_h="center")
        _write(ws, row, 9, f"Saldo: ${r['saldo_total']:,.0f}", fill_hex=COLOR_TOTAL,
               align_h="center", bold=True)
        ws.cell(row=row, column=10).fill = _fill(COLOR_TOTAL)
        row += 1

        fugadas_tot = int(r["fugadas"])
        fuga_pct    = _pct(fugadas_tot, tot)
        _write(ws, row, 1, f"Fuga (no aparecen en sig. {CHURN_WINDOW} camps.)", bold=True, fill_hex=COLOR_MORA3)
        _write(ws, row, 2, fugadas_tot, fill_hex=COLOR_MORA3, align_h="center", bold=True)
        _write(ws, row, 3, f"{fuga_pct}% del total", fill_hex=COLOR_MORA3, align_h="center")
        for col_e in range(4, 11):
            ws.cell(row=row, column=col_e).fill = _fill(COLOR_MORA3)
        row += 1

        # Encabezado tabla de mora (incluye Inactiva)
        headers = ["Nivel de Mora", "Total", "%", "Nuevas", "% Nuevas",
                   "Retenidas", "% Ret.", "Fugadas", "% Fuga", "Saldo"]
        for ci, h in enumerate(headers, 1):
            _header(ws, row, ci, h, color="374151")
        row += 1

        for mora in MORA_LEVELS:
            m_tot  = int(r.get(f"{mora}_total", 0))
            m_nue  = int(r.get(f"{mora}_nuevas", 0))
            m_ret  = int(r.get(f"{mora}_retenidas", 0))
            m_fug  = int(r.get(f"{mora}_fugadas", 0))
            m_sld  = r.get(f"{mora}_saldo", 0)
            bg     = mora_colors_hex.get(mora, COLOR_WHITE)
            vals   = [
                mora, m_tot, f"{_pct(m_tot, tot)}%",
                m_nue, f"{_pct(m_nue, m_tot)}%",
                m_ret, f"{_pct(m_ret, m_tot)}%",
                m_fug, f"{_pct(m_fug, m_tot)}%",
            ]
            for ci, v in enumerate(vals, 1):
                _write(ws, row, ci, v, fill_hex=bg, align_h="center" if ci > 1 else "left")
            _write(ws, row, 10, m_sld, fill_hex=bg, align_h="right", num_format='"$"#,##0')
            row += 1

        # Totales mora
        _write(ws, row, 1, "TOTAL", bold=True, fill_hex=COLOR_TOTAL)
        for ci in range(2, 11):
            ws.cell(row=row, column=ci).fill = _fill(COLOR_TOTAL)
        row += 1

        # ── Bloque de Inactivas: transición desde campaña anterior ──────
        fi_total = int(r.get("from_inactiva_total", 0))
        if fi_total > 0:
            _write(ws, row, 1,
                   "↳ Inactivas camp. anterior que reaparecen",
                   bold=True, fill_hex=COLOR_INACTIVA, wrap=True)
            _write(ws, row, 2, fi_total, fill_hex=COLOR_INACTIVA, align_h="center", bold=True)
            for col_e in range(3, 11):
                ws.cell(row=row, column=col_e).fill = _fill(COLOR_INACTIVA)
            row += 1

            # Detalle por estado destino
            for mora in MORA_LEVELS:
                cnt = int(r.get(f"from_inactiva_to_{mora}", 0))
                if cnt == 0:
                    continue
                m_tot_dest = int(r.get(f"{mora}_total", 0))
                pct_of_mora = _pct(cnt, m_tot_dest)
                pct_of_inac = _pct(cnt, fi_total)
                bg = mora_colors_hex.get(mora, COLOR_WHITE)
                _write(ws, row, 1, f"   → {mora}", fill_hex=bg)
                _write(ws, row, 2, cnt, fill_hex=bg, align_h="center")
                _write(ws, row, 3, f"{pct_of_inac}% de Inactivas", fill_hex=bg, align_h="center")
                _write(ws, row, 4, f"{pct_of_mora}% del total {mora}", fill_hex=bg, align_h="center")
                for col_e in range(5, 11):
                    ws.cell(row=row, column=col_e).fill = _fill(bg)
                row += 1

        row += 1  # espacio entre campañas

    ws.freeze_panes = "A3"

# ── Pestaña 2: Evolución de Mora ─────────────────────────────────────────────
def write_evolucion(wb: Workbook, data: dict):
    ws = wb.create_sheet("Evolución de Mora")
    ws.sheet_view.showGridLines = False
    trans_df = data["transitions"]
    camps    = data["camps"]
    labels   = data["labels"]

    n_cols = len(MORA_LEVELS) + 2  # origen label + mora cols + total

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"].value     = "MATRIZ DE TRANSICIÓN DE MORA ENTRE CAMPAÑAS  (incluye estado Inactiva)"
    ws["A1"].font      = Font(bold=True, size=13, color=COLOR_WHITE)
    ws["A1"].fill      = _fill(COLOR_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    if trans_df.empty:
        ws["A3"] = "Sin datos de transición disponibles."
        return

    row = 3
    mora_colors_hex = {
        "Inactiva": COLOR_INACTIVA,
        "Mora 1":   COLOR_MORA1,
        "Mora 2":   COLOR_MORA2,
        "Mora 3":   COLOR_MORA3,
    }

    for i in range(len(camps) - 1):
        c1, c2 = camps[i], camps[i + 1]
        sub = trans_df[(trans_df["de"] == labels[c1]) & (trans_df["a"] == labels[c2])]
        if sub.empty:
            continue

        # Pivot con todos los estados (incluye Inactiva)
        pivot = sub.pivot_table(index="origen", columns="destino",
                                values="cuentas", fill_value=0, aggfunc="sum")
        pivot = pivot.reindex(index=MORA_LEVELS, columns=MORA_LEVELS, fill_value=0)

        # Encabezado transición
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        cell = ws.cell(row=row, column=1,
                       value=f"  {labels[c1]} → {labels[c2]}   (damas presentes en ambas campañas)")
        cell.font      = Font(bold=True, size=11, color=COLOR_WHITE)
        cell.fill      = _fill(COLOR_SUBHEAD)
        cell.alignment = _align(h="left")
        ws.row_dimensions[row].height = 20
        row += 1

        # Instrucción
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        ws.cell(row=row, column=1).value = \
            "Filas = estado en campaña origen  |  Columnas = estado en campaña destino"
        ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="64748B")
        row += 1

        # Cabecera columnas
        _write(ws, row, 1, "Origen \\ Destino", bold=True, fill_hex="374151",
               font_color=COLOR_WHITE, align_h="center")
        for ci, m in enumerate(MORA_LEVELS, 2):
            _write(ws, row, ci, m, bold=True, fill_hex="374151",
                   font_color=COLOR_WHITE, align_h="center")
        _write(ws, row, n_cols, "TOTAL ORIGEN", bold=True,
               fill_hex="374151", font_color=COLOR_WHITE, align_h="center")
        row += 1

        for mora_orig in MORA_LEVELS:
            bg = mora_colors_hex.get(mora_orig, COLOR_WHITE)
            _write(ws, row, 1, mora_orig, bold=True, fill_hex=bg)
            row_total = 0
            for ci, mora_dest in enumerate(MORA_LEVELS, 2):
                val = int(pivot.loc[mora_orig, mora_dest]) \
                      if mora_orig in pivot.index and mora_dest in pivot.columns else 0
                row_total += val
                _write(ws, row, ci, val, fill_hex=bg, align_h="center")
            _write(ws, row, n_cols, row_total, bold=True,
                   fill_hex=COLOR_TOTAL, align_h="center")
            row += 1

        # Fila totales columna
        _write(ws, row, 1, "TOTAL DESTINO", bold=True, fill_hex=COLOR_TOTAL)
        for ci, mora_dest in enumerate(MORA_LEVELS, 2):
            col_total = int(pivot[mora_dest].sum()) if mora_dest in pivot.columns else 0
            _write(ws, row, ci, col_total, bold=True, fill_hex=COLOR_TOTAL, align_h="center")
        _write(ws, row, n_cols,
               int(pivot.values.sum()), bold=True, fill_hex=COLOR_TOTAL, align_h="center")
        row += 3

    # Ajustar anchos
    for col in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

# ── Pestaña 3: Gráficas ───────────────────────────────────────────────────────
def write_graficas(wb: Workbook, data: dict):
    ws = wb.create_sheet("Gráficas")
    ws.sheet_view.showGridLines = False
    summary = data["summary"]

    ws.merge_cells("A1:P1")
    ws["A1"].value     = "GRÁFICAS — EVOLUCIÓN DE CARTERA POR CAMPAÑA"
    ws["A1"].font      = Font(bold=True, size=13, color=COLOR_WHITE)
    ws["A1"].fill      = _fill(COLOR_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    # ── Tabla de datos para gráficas ──────────────────────────────────
    data_row = 3
    ws.cell(row=data_row, column=1).value = "Campaña"
    headers = ["Total", "Nuevas", "Retenidas", "Fugadas",
               "Inactiva", "Mora 1", "Mora 2", "Mora 3"]
    for ci, h in enumerate(headers, 2):
        ws.cell(row=data_row, column=ci).value = h
        ws.cell(row=data_row, column=ci).font = Font(bold=True)

    for ri, (_, r) in enumerate(summary.iterrows(), 1):
        ws.cell(row=data_row + ri, column=1).value = r["camp_label"]
        for ci, key in enumerate(["total", "nuevas", "retenidas", "fugadas",
                                   "Inactiva_total", "Mora 1_total",
                                   "Mora 2_total", "Mora 3_total"], 2):
            ws.cell(row=data_row + ri, column=ci).value = int(r.get(key, 0))

    # Columna adicional: cuentas que vienen de Inactiva del período anterior
    fi_col = 11
    ws.cell(row=data_row, column=fi_col).value = "De Inactiva→Mora"
    ws.cell(row=data_row, column=fi_col).font  = Font(bold=True)
    for ri, (_, r) in enumerate(summary.iterrows(), 1):
        ws.cell(row=data_row + ri, column=fi_col).value = int(r.get("from_inactiva_total", 0))

    n_camps  = len(summary)
    data_end = data_row + n_camps

    # ── Gráfica 1: Barras apiladas Nuevas / Retenidas / Fugadas ───────
    bar = BarChart()
    bar.type     = "col"
    bar.grouping = "stacked"
    bar.title    = "Composición por Campaña: Nuevas, Retenidas y Fugadas"
    bar.y_axis.title = "Número de Damas"
    bar.x_axis.title = "Campaña"
    bar.width, bar.height = 22, 14

    cats     = Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_end)
    data_ref = Reference(ws, min_col=2, max_col=4, min_row=data_row, max_row=data_end)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats)
    for i, clr in enumerate(["3B82F6", "22C55E", "EF4444"]):
        bar.series[i].graphicalProperties.solidFill = clr
    ws.add_chart(bar, "A3")

    # ── Gráfica 2: Líneas evolución de mora (incluye Inactiva) ────────
    line = LineChart()
    line.title        = "Evolución del Volumen por Estado (Inactiva + Moras)"
    line.y_axis.title = "Número de Damas"
    line.x_axis.title = "Campaña"
    line.width, line.height = 22, 14
    line.style = 10

    # cols 6-9 = Inactiva, Mora1, Mora2, Mora3
    data_ref2 = Reference(ws, min_col=6, max_col=9, min_row=data_row, max_row=data_end)
    line.add_data(data_ref2, titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_end))
    for i, clr in enumerate(["94A3B8", "F59E0B", "F97316", "EF4444"]):
        line.series[i].graphicalProperties.line.solidFill = clr
        line.series[i].graphicalProperties.line.width = 20000
    ws.add_chart(line, "A22")

    # ── Gráfica 3: Barras — migración desde Inactiva ──────────────────
    bar3 = BarChart()
    bar3.type         = "col"
    bar3.title        = "Cuentas que migran desde estado Inactiva (campaña anterior)"
    bar3.y_axis.title = "Número de Damas"
    bar3.x_axis.title = "Campaña"
    bar3.width, bar3.height = 18, 14
    data_ref4 = Reference(ws, min_col=fi_col, min_row=data_row, max_row=data_end)
    bar3.add_data(data_ref4, titles_from_data=True)
    bar3.set_categories(Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_end))
    bar3.series[0].graphicalProperties.solidFill = "6366F1"
    ws.add_chart(bar3, "M3")

    # ── Gráfica 4: Barras horizontales — % nuevas por campaña ─────────
    pct_col = 12
    ws.cell(row=data_row, column=pct_col).value = "% Nuevas"
    ws.cell(row=data_row, column=pct_col).font  = Font(bold=True)
    for ri, (_, r) in enumerate(summary.iterrows(), 1):
        pct = _pct(r["nuevas"], r["total"])
        ws.cell(row=data_row + ri, column=pct_col).value = pct / 100
        ws.cell(row=data_row + ri, column=pct_col).number_format = "0.0%"

    bar2 = BarChart()
    bar2.type         = "bar"
    bar2.title        = "% de Cuentas Nuevas por Campaña"
    bar2.y_axis.title = "Campaña"
    bar2.x_axis.title = "% Nuevas"
    bar2.width, bar2.height = 18, 14
    data_ref3 = Reference(ws, min_col=pct_col, min_row=data_row, max_row=data_end)
    bar2.add_data(data_ref3, titles_from_data=True)
    bar2.set_categories(Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_end))
    bar2.series[0].graphicalProperties.solidFill = "0EA5E9"
    ws.add_chart(bar2, "M22")

# ── Pestaña 4: Detalle de Fuga ────────────────────────────────────────────────
def write_fuga(wb: Workbook, data: dict):
    ws = wb.create_sheet("Detalle de Fuga")
    ws.sheet_view.showGridLines = False
    exits_df = data["exits_df"] if "exits_df" in data else data.get("exits", pd.DataFrame())

    ws.merge_cells("A1:D1")
    ws["A1"].value     = "DETALLE DE CUENTAS QUE SALIERON DEFINITIVAMENTE DEL FLUJO"
    ws["A1"].font      = Font(bold=True, size=13, color=COLOR_WHITE)
    ws["A1"].fill      = _fill(COLOR_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    ws.cell(row=3, column=1).value = "Total cuentas que no vuelven a aparecer:"
    ws.cell(row=3, column=1).font  = Font(bold=True)
    ws.cell(row=3, column=2).value = len(exits_df)
    ws.cell(row=3, column=2).font  = Font(bold=True, color="EF4444")

    ws.cell(row=4, column=1).value = "Campaña de salida"
    ws.cell(row=4, column=1).font  = Font(bold=True)

    row = 5
    mora_colors_hex = {
        "Inactiva": COLOR_INACTIVA,
        "Mora 1":   COLOR_MORA1,
        "Mora 2":   COLOR_MORA2,
        "Mora 3":   COLOR_MORA3,
    }

    if not exits_df.empty:
        grp = exits_df.groupby(["UltimaCampaña", "MoraAlSalir"]).agg(
            Cuentas=("NoDama", "count"),
            Saldo=("SaldoAlSalir", "sum")
        ).reset_index()

        hdrs = ["Última Campaña", "Estado al Salir", "Cuentas", "Saldo Total"]
        for ci, h in enumerate(hdrs, 1):
            _header(ws, row, ci, h)
        row += 1

        for _, r in grp.iterrows():
            bg = mora_colors_hex.get(r["MoraAlSalir"], COLOR_WHITE)
            _write(ws, row, 1, r["UltimaCampaña"], fill_hex=bg, align_h="center")
            _write(ws, row, 2, r["MoraAlSalir"],   fill_hex=bg)
            _write(ws, row, 3, int(r["Cuentas"]),  fill_hex=bg, align_h="center")
            _write(ws, row, 4, r["Saldo"],         fill_hex=bg, align_h="right",
                   num_format='"$"#,##0.00')
            row += 1

        _write(ws, row, 1, "TOTAL", bold=True, fill_hex=COLOR_TOTAL)
        _write(ws, row, 3, int(grp["Cuentas"].sum()), bold=True,
               fill_hex=COLOR_TOTAL, align_h="center")
        _write(ws, row, 4, grp["Saldo"].sum(), bold=True,
               fill_hex=COLOR_TOTAL, align_h="right", num_format='"$"#,##0.00')
        row += 3

        ws.cell(row=row, column=1).value = "LISTADO DE CUENTAS (para estrategia de recuperación)"
        ws.cell(row=row, column=1).font  = Font(bold=True, size=11)
        row += 1
        hdrs2 = ["NoDama", "Última Campaña", "Estado al Salir", "Saldo al Salir"]
        for ci, h in enumerate(hdrs2, 1):
            _header(ws, row, ci, h)
        row += 1
        for _, r in exits_df.head(5000).iterrows():
            bg = mora_colors_hex.get(r["MoraAlSalir"], COLOR_WHITE)
            _write(ws, row, 1, r["NoDama"],        fill_hex=bg)
            _write(ws, row, 2, r["UltimaCampaña"], fill_hex=bg, align_h="center")
            _write(ws, row, 3, r["MoraAlSalir"],   fill_hex=bg)
            _write(ws, row, 4, r["SaldoAlSalir"],  fill_hex=bg, align_h="right",
                   num_format='"$"#,##0.00')
            row += 1

    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 22

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("📂 Cargando datos...")
    df = load_data()
    print(f"   {len(df):,} registros · {df['NoDama'].nunique():,} damas únicas · "
          f"{df['Campania'].nunique()} campañas")

    # Distribución de estados incluyendo Inactiva
    print("\n📊 Distribución de estados (post-reclasificación):")
    dist = df["Moras"].value_counts()
    for estado, cnt in dist.items():
        print(f"   {estado:<12}: {cnt:>8,}  ({_pct(cnt, len(df)):.1f}%)")

    print("\n🔍 Calculando cohortes y métricas...")
    data = build_analysis(df)
    data["exits_df"] = data.pop("exits")

    print("📊 Generando Excel ejecutivo...")
    wb = Workbook()
    wb.remove(wb.active)

    write_resumen(wb, data)
    write_evolucion(wb, data)
    write_graficas(wb, data)
    write_fuga(wb, data)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Archivo guardado en: {OUTPUT_FILE}")
    print(f"   Pestañas: {[s.title for s in wb.worksheets]}")

    # Resumen en consola
    summary = data["summary"]
    print("\n📋 RESUMEN RÁPIDO (incluye Inactivas):")
    print(f"{'Camp.':<6} {'Total':>8} {'Nuevas':>8} {'%Nue':>6} {'Retenidas':>10} "
          f"{'Fugadas':>8} {'Inactiva':>10} {'Mora1':>7} {'Mora2':>7} {'Mora3':>7} {'De Inac→':>9}")
    print("-" * 90)
    for _, r in summary.iterrows():
        tot = int(r["total"])
        print(
            f"{r['camp_label']:<6} {tot:>8,} {int(r['nuevas']):>8,} "
            f"{_pct(r['nuevas'], tot):>5.1f}% {int(r['retenidas']):>10,} "
            f"{int(r['fugadas']):>8,} {int(r.get('Inactiva_total', 0)):>10,} "
            f"{int(r.get('Mora 1_total', 0)):>7,} {int(r.get('Mora 2_total', 0)):>7,} "
            f"{int(r.get('Mora 3_total', 0)):>7,} {int(r.get('from_inactiva_total', 0)):>9,}"
        )

    # Detalle de transiciones Inactiva → Mora
    print("\n🔄 TRANSICIÓN INACTIVAS → ESTADOS (campaña anterior → actual):")
    print(f"{'Camp.':6} {'De Inac→':>9} {'→Inac':>8} {'→Mora1':>8} {'→Mora2':>8} {'→Mora3':>8}")
    print("-" * 55)
    for _, r in summary.iterrows():
        fi = int(r.get("from_inactiva_total", 0))
        if fi == 0:
            continue
        print(
            f"{r['camp_label']:<6} {fi:>9,} "
            f"{int(r.get('from_inactiva_to_Inactiva', 0)):>8,} "
            f"{int(r.get('from_inactiva_to_Mora 1', 0)):>8,} "
            f"{int(r.get('from_inactiva_to_Mora 2', 0)):>8,} "
            f"{int(r.get('from_inactiva_to_Mora 3', 0)):>8,}"
        )

if __name__ == "__main__":
    main()
