"""
tracking_pendientes.py
Comprehensive campaign tracking analysis — generates polished executive Excel report.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── Paths ──────────────────────────────────────────────────────────────────────
INPUT_FILE  = "/root/.claude/uploads/48129c5c-b81b-412d-8c25-6859dcfca1e5/ba7a1adc-GENERAL_CARTERA_MORAS_CAMPA_A_1__10_2.xlsx"
OUTPUT_FILE = "/home/user/Arabela/reporte_tracking_pendientes.xlsx"

# ── Constants ──────────────────────────────────────────────────────────────────
MORA_LEVELS  = ["Inactiva", "Mora 1", "Mora 2", "Mora 3"]
CHURN_WINDOW = 3

# ── Color palette ─────────────────────────────────────────────────────────────
COLOR_HEADER   = "1A3C6E"
COLOR_SUBHEAD  = "2563EB"
COLOR_INACTIVA = "E2E8F0"
COLOR_MORA1    = "FEF9C3"
COLOR_MORA2    = "FFEDD5"
COLOR_MORA3    = "FEE2E2"
COLOR_TOTAL    = "DBEAFE"
COLOR_ALT      = "F1F5F9"
COLOR_WHITE    = "FFFFFF"
COLOR_SUCCESS  = "DCFCE7"
COLOR_POOL     = "F3E8FF"

MORA_COLORS = {
    "Inactiva": COLOR_INACTIVA,
    "Mora 1":   COLOR_MORA1,
    "Mora 2":   COLOR_MORA2,
    "Mora 3":   COLOR_MORA3,
}

THIN_BORDER_COLOR = "CBD5E1"


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=11, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic)

def make_align(horizontal="center", vertical="center", wrap=False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def thin_border(top=True, bottom=True, left=True, right=True) -> Border:
    side = Side(style="thin", color=THIN_BORDER_COLOR)
    no   = Side(style=None)
    return Border(
        top    = side if top    else no,
        bottom = side if bottom else no,
        left   = side if left   else no,
        right  = side if right  else no,
    )

def pct(num, den, digits=1):
    if den == 0:
        return 0.0
    return round(num / den * 100, digits)

def fmt_num(n) -> str:
    return f"{int(n):,}"

def fmt_pct(v) -> str:
    return f"{v:.1f}%"

def fmt_currency(v) -> str:
    return f"${v:,.2f}"

def style_header_cell(ws, row, col, value, size=12, merge_to=None):
    """Write a primary header cell (dark blue bg, white bold text)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = make_fill(COLOR_HEADER)
    cell.font   = make_font(bold=True, color=COLOR_WHITE, size=size)
    cell.alignment = make_align(horizontal="center", wrap=True)
    cell.border = thin_border()
    if merge_to:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=merge_to,
        )
    return cell

def style_subheader_cell(ws, row, col, value, size=11, merge_to=None):
    """Write a sub-header cell (medium blue bg, white bold text)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = make_fill(COLOR_SUBHEAD)
    cell.font   = make_font(bold=True, color=COLOR_WHITE, size=size)
    cell.alignment = make_align(horizontal="center", wrap=True)
    cell.border = thin_border()
    if merge_to:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=merge_to,
        )
    return cell

def style_data_cell(ws, row, col, value, alt=False, fill_hex=None,
                    bold=False, number_format=None, align="right"):
    cell = ws.cell(row=row, column=col, value=value)
    bg = fill_hex if fill_hex else (COLOR_ALT if alt else COLOR_WHITE)
    cell.fill   = make_fill(bg)
    cell.font   = make_font(bold=bold)
    cell.alignment = make_align(horizontal=align)
    cell.border = thin_border()
    if number_format:
        cell.number_format = number_format
    return cell

def style_label_cell(ws, row, col, value, fill_hex=COLOR_WHITE, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = make_fill(fill_hex)
    cell.font   = make_font(bold=bold)
    cell.alignment = make_align(horizontal="left")
    cell.border = thin_border()
    return cell

def set_col_widths(ws, widths: dict):
    """widths = {col_letter_or_int: width}"""
    for col, w in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def autofit_cols(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_data() -> pd.DataFrame:
    print("Loading data …")
    df = pd.read_excel(INPUT_FILE, sheet_name="Consolidado")
    df.columns = [c.strip() for c in df.columns]
    df["NoDama"]   = df["NoDama"].astype(str).str.strip()
    df["Campania"] = df["Campania"].astype(str).str.strip().str.upper()
    df["Moras"]    = df["Moras"].astype(str).str.strip()
    df["SaldoDama"] = pd.to_numeric(df["SaldoDama"], errors="coerce").fillna(0)
    df["_camp_n"]  = (
        df["Campania"]
        .str.replace("C", "", regex=False)
        .apply(lambda x: int(x) if x.isdigit() else None)
    )
    df = df.dropna(subset=["_camp_n"]).copy()
    df["_camp_n"]  = df["_camp_n"].astype(int)
    if "IdSituacion" in df.columns:
        df.loc[df["IdSituacion"].astype(str).str.strip() == "0", "Moras"] = "Inactiva"
    print(f"  Rows: {len(df):,}  |  Unique NoDamas: {df['NoDama'].nunique():,}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def build_analysis(df: pd.DataFrame) -> dict:
    print("Building analysis …")
    camps      = sorted(df["_camp_n"].unique())
    pool       = set(df["NoDama"].unique())
    sets       = {c: set(df[df["_camp_n"] == c]["NoDama"]) for c in camps}
    first_camp = df.groupby("NoDama")["_camp_n"].min().to_dict()

    mora_sets: dict = {}
    for c in camps:
        df_c = df[df["_camp_n"] == c]
        for mora in MORA_LEVELS:
            mora_sets[(c, mora)] = set(df_c[df_c["Moras"] == mora]["NoDama"])

    results = []
    for i, c in enumerate(camps):
        prev_c  = camps[i - 1] if i > 0 else None
        prev2_c = camps[i - 2] if i > 1 else None
        total_set = sets[c]
        nuevas    = {d for d in total_set if first_camp.get(d) == c}
        de_anterior  = (total_set & sets[prev_c])  if prev_c  else set()
        persistentes = (total_set & sets[prev_c] & sets[prev2_c]) if (prev_c and prev2_c) else set()
        next_c   = camps[i + 1] if i < len(camps) - 1 else None
        salen    = (total_set - sets[next_c]) if next_c else set()
        future   = [cx for cx in camps if cx > c and cx <= c + CHURN_WINDOW]
        fugadas  = (
            (total_set - set().union(*[sets[cx] for cx in future]))
            if future else set()
        )

        row = {
            "camp_n":      c,
            "camp_label":  f"C{c}",
            "pool_total":  len(pool),
            "total":       len(total_set),
            "pct_pool":    round(len(total_set) / len(pool) * 100, 1),
            "nuevas":      len(nuevas),
            "de_anterior": len(de_anterior),
            "persistentes":len(persistentes),
            "salen_siguiente": len(salen),
            "fugadas":     len(fugadas),
            "saldo_total": df[df["_camp_n"] == c]["SaldoDama"].sum(),
        }

        for mora in MORA_LEVELS:
            ids_m = mora_sets[(c, mora)]
            row[f"{mora}_total"]       = len(ids_m)
            row[f"{mora}_nuevas"]      = len(ids_m & nuevas)
            row[f"{mora}_de_anterior"] = len(ids_m & de_anterior)
            row[f"{mora}_persistentes"]= len(ids_m & persistentes)
            row[f"{mora}_salen"]       = len(ids_m & salen)
            row[f"{mora}_fugadas"]     = len(ids_m & fugadas)
            row[f"{mora}_saldo"]       = float(
                df[(df["_camp_n"] == c) & (df["Moras"] == mora)]["SaldoDama"].sum()
            )

        if prev_c:
            inac_prev = mora_sets[(prev_c, "Inactiva")]
            row["fi_total"] = len(inac_prev & total_set)
            for mora in MORA_LEVELS:
                row[f"fi_to_{mora}"] = len(inac_prev & mora_sets[(c, mora)])
        else:
            row["fi_total"] = 0
            for mora in MORA_LEVELS:
                row[f"fi_to_{mora}"] = 0

        results.append(row)

    summary_df = pd.DataFrame(results)

    # Transition matrix
    transitions = []
    for i in range(len(camps) - 1):
        c1, c2 = camps[i], camps[i + 1]
        shared = sets[c1] & sets[c2]
        df1 = (
            df[(df["_camp_n"] == c1) & df["NoDama"].isin(shared)]
            .drop_duplicates("NoDama").set_index("NoDama")["Moras"]
        )
        df2 = (
            df[(df["_camp_n"] == c2) & df["NoDama"].isin(shared)]
            .drop_duplicates("NoDama").set_index("NoDama")["Moras"]
        )
        joined = df1.to_frame("origen").join(df2.to_frame("destino"), how="inner")
        for (orig, dest), cnt in joined.groupby(["origen", "destino"]).size().items():
            transitions.append({
                "de": f"C{c1}", "a": f"C{c2}",
                "origen": orig, "destino": dest, "cuentas": int(cnt),
            })
    trans_df = pd.DataFrame(transitions) if transitions else pd.DataFrame()

    # Exit detail
    last_camp = df.groupby("NoDama")["_camp_n"].max().to_dict()
    max_c = max(camps)
    exits = []
    for dama, lc in last_camp.items():
        if lc < max_c:
            sub    = df[(df["NoDama"] == dama) & (df["_camp_n"] == lc)]
            mora_e = sub["Moras"].iloc[0]   if len(sub) > 0 else "N/A"
            saldo_e= float(sub["SaldoDama"].iloc[0]) if len(sub) > 0 else 0.0
            exits.append({
                "NoDama":        dama,
                "UltimaCampaña": f"C{lc}",
                "MoraAlSalir":   mora_e,
                "SaldoAlSalir":  saldo_e,
            })
    exits_df = pd.DataFrame(exits)

    return {
        "summary":     summary_df,
        "transitions": trans_df,
        "exits":       exits_df,
        "camps":       camps,
        "pool":        pool,
    }


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Resumen Ejecutivo
# ══════════════════════════════════════════════════════════════════════════════

def write_resumen_ejecutivo(wb: Workbook, data: dict):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False

    summary = data["summary"]
    pool_size = len(data["pool"])

    # ── Main title ─────────────────────────────────────────────────────────
    style_header_cell(ws, 1, 1,
        "TRACKING DE CARTERA — ANÁLISIS ENTRE CAMPAÑAS",
        size=14, merge_to=16)
    ws.row_dimensions[1].height = 30

    style_subheader_cell(ws, 2, 1,
        f"{pool_size:,} pendientes de pago en seguimiento  |  C1–C10  |  Ventana de fuga: {CHURN_WINDOW} campañas",
        size=11, merge_to=16)
    ws.row_dimensions[2].height = 22

    # ── Column headers ─────────────────────────────────────────────────────
    headers = [
        "Campaña", "Pendientes", "% del Pool", "Nuevas", "% Nuevas",
        "De Camp. Anterior", "% Retención", "Persistentes\n(2+ camps)",
        "Salen Sig.\nCamp.", "Fugadas\n(3c)", "% Fuga",
        "Inactiva", "Mora 1", "Mora 2", "Mora 3", "Saldo Total",
    ]
    for col_i, h in enumerate(headers, start=1):
        style_subheader_cell(ws, 4, col_i, h, size=10)
    ws.row_dimensions[4].height = 36

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_i, row in summary.iterrows():
        r = 5 + row_i
        alt = row_i % 2 == 1
        bg  = COLOR_ALT if alt else COLOR_WHITE

        style_data_cell(ws, r, 1,  row["camp_label"],    fill_hex=bg, bold=True, align="center")
        style_data_cell(ws, r, 2,  row["total"],          fill_hex=bg, number_format="#,##0")
        style_data_cell(ws, r, 3,  row["pct_pool"],       fill_hex=bg, number_format='0.0"%"')
        style_data_cell(ws, r, 4,  row["nuevas"],         fill_hex=COLOR_SUCCESS, number_format="#,##0")
        style_data_cell(ws, r, 5,  pct(row["nuevas"], row["total"]),
                        fill_hex=COLOR_SUCCESS, number_format='0.0"%"')
        style_data_cell(ws, r, 6,  row["de_anterior"],    fill_hex=bg, number_format="#,##0")
        style_data_cell(ws, r, 7,  pct(row["de_anterior"],
                                       summary.iloc[row_i - 1]["total"] if row_i > 0 else 0),
                        fill_hex=bg, number_format='0.0"%"')
        style_data_cell(ws, r, 8,  row["persistentes"],   fill_hex=bg, number_format="#,##0")
        style_data_cell(ws, r, 9,  row["salen_siguiente"],fill_hex=bg, number_format="#,##0")
        style_data_cell(ws, r, 10, row["fugadas"],         fill_hex=bg, number_format="#,##0")
        style_data_cell(ws, r, 11, pct(row["fugadas"], row["total"]),
                        fill_hex=bg, number_format='0.0"%"')
        style_data_cell(ws, r, 12, row["Inactiva_total"],  fill_hex=COLOR_INACTIVA, number_format="#,##0")
        style_data_cell(ws, r, 13, row["Mora 1_total"],    fill_hex=COLOR_MORA1,    number_format="#,##0")
        style_data_cell(ws, r, 14, row["Mora 2_total"],    fill_hex=COLOR_MORA2,    number_format="#,##0")
        style_data_cell(ws, r, 15, row["Mora 3_total"],    fill_hex=COLOR_MORA3,    number_format="#,##0")
        style_data_cell(ws, r, 16, row["saldo_total"],     fill_hex=bg,
                        number_format='"$"#,##0.00')

    # ── Key metrics section ────────────────────────────────────────────────
    last_data_row = 5 + len(summary)
    km_start = last_data_row + 3

    style_header_cell(ws, km_start, 1, "MÉTRICAS CLAVE", size=12, merge_to=8)
    ws.row_dimensions[km_start].height = 24

    # Avg retention (C2–C10: de_anterior / prev_total)
    retenciones = []
    for i in range(1, len(summary)):
        prev_total = summary.iloc[i - 1]["total"]
        if prev_total > 0:
            retenciones.append(summary.iloc[i]["de_anterior"] / prev_total * 100)
    avg_ret = round(np.mean(retenciones), 1) if retenciones else 0.0

    # Campaign with highest fugadas %
    summary["_pct_fuga"] = summary.apply(
        lambda r: pct(r["fugadas"], r["total"]), axis=1
    )
    max_fuga_idx = summary["_pct_fuga"].idxmax()
    max_fuga_camp = summary.loc[max_fuga_idx, "camp_label"]
    max_fuga_val  = summary.loc[max_fuga_idx, "_pct_fuga"]

    # Campaign with most nuevas
    max_new_idx   = summary["nuevas"].idxmax()
    max_new_camp  = summary.loc[max_new_idx, "camp_label"]
    max_new_val   = summary.loc[max_new_idx, "nuevas"]

    metrics = [
        ("Total NoDamas únicas en seguimiento:",         f"{pool_size:,}"),
        ("Tasa promedio de retención entre campañas:",   f"{avg_ret:.1f}%"),
        (f"Campaña con mayor fuga:",                      f"{max_fuga_camp}  ({max_fuga_val:.1f}%)"),
        ("Campaña con más cuentas nuevas:",               f"{max_new_camp}  ({max_new_val:,} nuevas)"),
    ]
    for mi, (label, value) in enumerate(metrics):
        r = km_start + 1 + mi
        style_label_cell(ws, r, 1, label, fill_hex=COLOR_TOTAL, bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        style_data_cell(ws, r, 6, value, fill_hex=COLOR_WHITE, align="left", bold=True)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)

    # Column widths
    widths = {1:12, 2:12, 3:10, 4:10, 5:10, 6:14, 7:12, 8:16,
              9:12, 10:12, 11:10, 12:10, 13:10, 14:10, 15:10, 16:16}
    set_col_widths(ws, widths)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Análisis por Campaña
# ══════════════════════════════════════════════════════════════════════════════

def write_analisis_por_campana(wb: Workbook, data: dict):
    ws = wb.create_sheet("Análisis por Campaña")
    ws.sheet_view.showGridLines = False

    summary = data["summary"]
    cur_row = 1

    # Sheet title
    style_header_cell(ws, cur_row, 1, "ANÁLISIS DETALLADO POR CAMPAÑA", size=13, merge_to=12)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 2

    for _, row in summary.iterrows():
        c      = row["camp_n"]
        clabel = row["camp_label"]

        # ── Campaign header ────────────────────────────────────────────────
        style_header_cell(ws, cur_row, 1,
            f"{'─'*5}  {clabel}  {'─'*5}",
            size=12, merge_to=12)
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        # ── KPI row ────────────────────────────────────────────────────────
        kpi_labels = ["Total", "Nuevas", "De Camp. Anterior",
                      "Persistentes", "Salen Sig.", "Fugadas", "Saldo Total"]
        kpi_vals   = [
            fmt_num(row["total"]),
            fmt_num(row["nuevas"]),
            fmt_num(row["de_anterior"]),
            fmt_num(row["persistentes"]),
            fmt_num(row["salen_siguiente"]),
            fmt_num(row["fugadas"]),
            fmt_currency(row["saldo_total"]),
        ]
        kpi_colors = [COLOR_TOTAL, COLOR_SUCCESS, COLOR_ALT,
                      COLOR_ALT, COLOR_ALT, COLOR_MORA2, COLOR_WHITE]
        for ki, (lbl, val, kc) in enumerate(zip(kpi_labels, kpi_vals, kpi_colors)):
            col = 1 + ki * 1
            cell_l = ws.cell(row=cur_row, column=col * 2 - 1, value=lbl)
            cell_l.fill = make_fill(kc)
            cell_l.font = make_font(bold=True, size=10)
            cell_l.alignment = make_align(horizontal="right")
            cell_l.border = thin_border()

            cell_v = ws.cell(row=cur_row, column=col * 2, value=val)
            cell_v.fill = make_fill(kc)
            cell_v.font = make_font(size=10)
            cell_v.alignment = make_align(horizontal="left")
            cell_v.border = thin_border()

        # blank leftover cols
        for bcol in range(15, 13, -1):
            ws.cell(row=cur_row, column=bcol).fill = make_fill(COLOR_WHITE)

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # ── Mora breakdown table ───────────────────────────────────────────
        mora_hdrs = [
            "Estado", "Total", "%",
            "Nuevas", "% Nuevas",
            "De Anterior", "% De Ant.",
            "Persistentes", "Salen", "Fugadas", "% Fuga", "Saldo",
        ]
        for ci, h in enumerate(mora_hdrs, start=1):
            style_subheader_cell(ws, cur_row, ci, h, size=10)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        for mora in MORA_LEVELS:
            mc    = MORA_COLORS[mora]
            mt    = row[f"{mora}_total"]
            mn    = row[f"{mora}_nuevas"]
            mda   = row[f"{mora}_de_anterior"]
            mper  = row[f"{mora}_persistentes"]
            msal  = row[f"{mora}_salen"]
            mfug  = row[f"{mora}_fugadas"]
            msaldo= row[f"{mora}_saldo"]

            vals = [
                mora, mt,
                pct(mt, row["total"]),
                mn,
                pct(mn, mt),
                mda,
                pct(mda, mt),
                mper, msal, mfug,
                pct(mfug, mt),
                msaldo,
            ]
            fmts = [
                None, "#,##0", '0.0"%"',
                "#,##0", '0.0"%"',
                "#,##0", '0.0"%"',
                "#,##0", "#,##0", "#,##0",
                '0.0"%"',
                '"$"#,##0.00',
            ]
            aligns = ["left"] + ["right"] * 11

            for ci, (v, fmt, al) in enumerate(zip(vals, fmts, aligns), start=1):
                cell = ws.cell(row=cur_row, column=ci, value=v)
                cell.fill = make_fill(mc)
                cell.font = make_font(bold=(ci == 1))
                cell.alignment = make_align(horizontal=al)
                cell.border = thin_border()
                if fmt:
                    cell.number_format = fmt

            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        # ── Inactiva transitions ───────────────────────────────────────────
        if row["fi_total"] > 0:
            cur_row += 1
            ti_cell = ws.cell(row=cur_row, column=1,
                value=f"Inactivas de campaña anterior que reaparecen: {fmt_num(row['fi_total'])}")
            ti_cell.fill = make_fill(COLOR_INACTIVA)
            ti_cell.font = make_font(bold=True, size=10)
            ti_cell.alignment = make_align(horizontal="left")
            ti_cell.border = thin_border()
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=8)
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

            for mora in MORA_LEVELS:
                val = row[f"fi_to_{mora}"]
                mc = MORA_COLORS[mora]
                label_cell = ws.cell(row=cur_row, column=1, value=f"  → {mora}:")
                label_cell.fill = make_fill(mc)
                label_cell.font = make_font(bold=True, size=10)
                label_cell.alignment = make_align(horizontal="left")
                label_cell.border = thin_border()
                ws.merge_cells(start_row=cur_row, start_column=1,
                               end_row=cur_row, end_column=3)

                val_cell = ws.cell(row=cur_row, column=4, value=val)
                val_cell.fill = make_fill(mc)
                val_cell.font = make_font(size=10)
                val_cell.alignment = make_align(horizontal="right")
                val_cell.border = thin_border()
                val_cell.number_format = "#,##0"
                ws.merge_cells(start_row=cur_row, start_column=4,
                               end_row=cur_row, end_column=5)
                ws.row_dimensions[cur_row].height = 15
                cur_row += 1

        cur_row += 2  # spacer between campaigns

    # Column widths
    widths = {1:14, 2:12, 3:10, 4:12, 5:10, 6:14, 7:12, 8:14, 9:12, 10:12, 11:10, 12:16}
    set_col_widths(ws, widths)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Análisis por Mora
# ══════════════════════════════════════════════════════════════════════════════

def write_analisis_por_mora(wb: Workbook, data: dict):
    ws = wb.create_sheet("Análisis por Mora")
    ws.sheet_view.showGridLines = False

    summary = data["summary"]
    camps   = data["camps"]
    cur_row = 1

    style_header_cell(ws, cur_row, 1, "ANÁLISIS POR ESTADO DE MORA", size=13, merge_to=10)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 2

    mora_cols = [
        "Campaña", "Total", "% del Pool Mora",
        "Nuevas", "% Nuevas",
        "De Anterior", "% Retención",
        "Fugadas", "% Fuga", "Saldo",
    ]

    for mora in MORA_LEVELS:
        mc = MORA_COLORS[mora]

        # Section header
        style_header_cell(ws, cur_row, 1,
            f"ESTADO: {mora.upper()}",
            size=12, merge_to=10)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Column headers
        for ci, h in enumerate(mora_cols, start=1):
            style_subheader_cell(ws, cur_row, ci, h, size=10)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        prev_total_mora = 0
        for ri, row in summary.iterrows():
            alt = ri % 2 == 0
            bg  = COLOR_ALT if alt else COLOR_WHITE
            mt  = row[f"{mora}_total"]
            mn  = row[f"{mora}_nuevas"]
            mda = row[f"{mora}_de_anterior"]
            mfug= row[f"{mora}_fugadas"]
            ms  = row[f"{mora}_saldo"]

            vals  = [
                row["camp_label"],
                mt,
                pct(mt, row["total"]),
                mn,
                pct(mn, mt),
                mda,
                pct(mda, prev_total_mora),
                mfug,
                pct(mfug, mt),
                ms,
            ]
            fmts  = [
                None, "#,##0", '0.0"%"',
                "#,##0", '0.0"%"',
                "#,##0", '0.0"%"',
                "#,##0", '0.0"%"',
                '"$"#,##0.00',
            ]
            aligns = ["center"] + ["right"] * 9

            for ci, (v, fmt, al) in enumerate(zip(vals, fmts, aligns), start=1):
                cell = ws.cell(row=cur_row, column=ci, value=v)
                cell.fill = make_fill(mc)
                cell.font = make_font()
                cell.alignment = make_align(horizontal=al)
                cell.border = thin_border()
                if fmt:
                    cell.number_format = fmt

            prev_total_mora = mt
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        cur_row += 2

    # ── Trend comparison table ─────────────────────────────────────────────
    style_header_cell(ws, cur_row, 1, "COMPARATIVO DE TENDENCIAS", size=12, merge_to=10)
    ws.row_dimensions[cur_row].height = 22
    cur_row += 1

    trend_hdrs = ["Campaña"] + MORA_LEVELS + ["Total"]
    for ci, h in enumerate(trend_hdrs, start=1):
        style_subheader_cell(ws, cur_row, ci, h, size=10)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    for ri, row in summary.iterrows():
        alt = ri % 2 == 0
        bg  = COLOR_ALT if alt else COLOR_WHITE
        style_data_cell(ws, cur_row, 1, row["camp_label"], fill_hex=bg, bold=True, align="center")
        for mi, mora in enumerate(MORA_LEVELS, start=2):
            style_data_cell(ws, cur_row, mi, row[f"{mora}_total"],
                            fill_hex=MORA_COLORS[mora], number_format="#,##0")
        style_data_cell(ws, cur_row, 6, row["total"],
                        fill_hex=COLOR_TOTAL, number_format="#,##0", bold=True)
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    widths = {1:14, 2:12, 3:14, 4:12, 5:14, 6:12, 7:14, 8:12, 9:12, 10:16}
    set_col_widths(ws, widths)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Tracking de Salidas
# ══════════════════════════════════════════════════════════════════════════════

def write_tracking_salidas(wb: Workbook, data: dict):
    ws = wb.create_sheet("Tracking de Salidas")
    ws.sheet_view.showGridLines = False

    exits_df = data["exits"]
    cur_row  = 1

    style_header_cell(ws, cur_row, 1,
        "CUENTAS QUE SALEN DEL FLUJO DE CAMPAÑAS",
        size=13, merge_to=8)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 2

    # ── Summary totals ─────────────────────────────────────────────────────
    total_exits = len(exits_df)
    style_subheader_cell(ws, cur_row, 1,
        f"Total que no aparecen en C10: {total_exits:,}",
        size=11, merge_to=8)
    ws.row_dimensions[cur_row].height = 20
    cur_row += 2

    # By last campaign
    style_subheader_cell(ws, cur_row, 1, "Por Última Campaña", size=10, merge_to=4)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    by_camp_hdrs = ["Última Campaña", "Cuentas", "% del Total", "Saldo Promedio"]
    for ci, h in enumerate(by_camp_hdrs, start=1):
        style_subheader_cell(ws, cur_row, ci, h, size=10)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    if not exits_df.empty:
        by_camp = exits_df.groupby("UltimaCampaña").agg(
            Count=("NoDama", "count"),
            AvgSaldo=("SaldoAlSalir", "mean"),
        ).reset_index().sort_values("UltimaCampaña")
        for ri, brow in by_camp.iterrows():
            alt = ri % 2 == 0
            bg  = COLOR_ALT if alt else COLOR_WHITE
            style_data_cell(ws, cur_row, 1, brow["UltimaCampaña"], fill_hex=bg, align="center")
            style_data_cell(ws, cur_row, 2, brow["Count"],   fill_hex=bg, number_format="#,##0")
            style_data_cell(ws, cur_row, 3, pct(brow["Count"], total_exits),
                            fill_hex=bg, number_format='0.0"%"')
            style_data_cell(ws, cur_row, 4, brow["AvgSaldo"],fill_hex=bg, number_format='"$"#,##0.00')
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

    cur_row += 2

    # By mora at exit
    style_subheader_cell(ws, cur_row, 1, "Por Mora al Salir", size=10, merge_to=4)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    mora_hdrs = ["Mora", "Cuentas", "%"]
    for ci, h in enumerate(mora_hdrs, start=1):
        style_subheader_cell(ws, cur_row, ci, h, size=10)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    if not exits_df.empty:
        by_mora = exits_df.groupby("MoraAlSalir").size().reset_index(name="Count")
        for ri, brow in by_mora.iterrows():
            mora  = brow["MoraAlSalir"]
            mc    = MORA_COLORS.get(mora, COLOR_WHITE)
            style_data_cell(ws, cur_row, 1, mora,          fill_hex=mc, align="left", bold=True)
            style_data_cell(ws, cur_row, 2, brow["Count"], fill_hex=mc, number_format="#,##0")
            style_data_cell(ws, cur_row, 3, pct(brow["Count"], total_exits),
                            fill_hex=mc, number_format='0.0"%"')
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

    cur_row += 2

    # Pivot: salidas por campaña y mora
    if not exits_df.empty:
        style_subheader_cell(ws, cur_row, 1,
            "Salidas por Campaña y Mora (pivot)", size=10, merge_to=8)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        pivot = (
            exits_df.groupby(["UltimaCampaña", "MoraAlSalir"])
            .size().unstack(fill_value=0).reset_index()
        )
        pivot_cols = list(pivot.columns)
        for ci, h in enumerate(pivot_cols, start=1):
            style_subheader_cell(ws, cur_row, ci, h, size=10)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        for ri, prow in pivot.iterrows():
            alt = ri % 2 == 0
            bg  = COLOR_ALT if alt else COLOR_WHITE
            for ci, col in enumerate(pivot_cols, start=1):
                v = prow[col]
                mc_use = bg if ci > 1 else bg
                if ci > 1:
                    col_name = pivot_cols[ci - 1]
                    mc_use = MORA_COLORS.get(col_name, bg)
                style_data_cell(ws, cur_row, ci, v,
                                fill_hex=mc_use,
                                number_format="#,##0" if ci > 1 else None,
                                align="center" if ci == 1 else "right")
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

    cur_row += 2

    # Listado detallado (first 5000)
    style_subheader_cell(ws, cur_row, 1,
        "Listado Detallado (primeras 5,000 cuentas)", size=10, merge_to=4)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    det_hdrs = ["NoDama", "Última Campaña", "Mora al Salir", "Saldo al Salir"]
    for ci, h in enumerate(det_hdrs, start=1):
        style_subheader_cell(ws, cur_row, ci, h, size=10)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    if not exits_df.empty:
        sample = exits_df.head(5000)
        for ri, drow in sample.iterrows():
            alt = ri % 2 == 0
            bg  = COLOR_ALT if alt else COLOR_WHITE
            mc  = MORA_COLORS.get(drow["MoraAlSalir"], bg)
            style_data_cell(ws, cur_row, 1, drow["NoDama"],        fill_hex=bg,  align="left")
            style_data_cell(ws, cur_row, 2, drow["UltimaCampaña"], fill_hex=bg,  align="center")
            style_data_cell(ws, cur_row, 3, drow["MoraAlSalir"],   fill_hex=mc,  align="left")
            style_data_cell(ws, cur_row, 4, drow["SaldoAlSalir"],  fill_hex=bg,
                            number_format='"$"#,##0.00')
            ws.row_dimensions[cur_row].height = 15
            cur_row += 1

    widths = {1:16, 2:16, 3:14, 4:16, 5:12, 6:12, 7:12, 8:14}
    set_col_widths(ws, widths)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Matriz de Transición
# ══════════════════════════════════════════════════════════════════════════════

def write_matriz_transicion(wb: Workbook, data: dict):
    ws = wb.create_sheet("Matriz de Transición")
    ws.sheet_view.showGridLines = False

    trans_df = data["transitions"]
    camps    = data["camps"]
    cur_row  = 1

    style_header_cell(ws, cur_row, 1,
        "MATRIZ DE TRANSICIÓN ENTRE CAMPAÑAS CONSECUTIVAS",
        size=13, merge_to=7)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 2

    if trans_df.empty:
        ws.cell(row=cur_row, column=1, value="Sin datos de transición.")
        return

    for i in range(len(camps) - 1):
        c1_label = f"C{camps[i]}"
        c2_label = f"C{camps[i+1]}"

        subset = trans_df[(trans_df["de"] == c1_label) & (trans_df["a"] == c2_label)]
        if subset.empty:
            continue

        # Section header
        style_header_cell(ws, cur_row, 1,
            f"Transición {c1_label} → {c2_label}",
            size=12, merge_to=7)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Pivot
        pivot = subset.pivot_table(
            index="origen", columns="destino", values="cuentas",
            aggfunc="sum", fill_value=0,
        )
        all_states = MORA_LEVELS
        for s in all_states:
            if s not in pivot.index:
                pivot.loc[s] = 0
            if s not in pivot.columns:
                pivot[s] = 0
        pivot = pivot.reindex(index=all_states, columns=all_states, fill_value=0)

        # Column headers: blank | destino states | TOTAL
        ws.cell(row=cur_row, column=1, value="Origen \\ Destino").fill = make_fill(COLOR_HEADER)
        ws.cell(row=cur_row, column=1).font   = make_font(bold=True, color=COLOR_WHITE, size=10)
        ws.cell(row=cur_row, column=1).alignment = make_align(horizontal="center")
        ws.cell(row=cur_row, column=1).border = thin_border()

        col_totals = {s: int(pivot[s].sum()) for s in all_states}
        grand_total = int(pivot.values.sum())

        for ci, dest_state in enumerate(all_states, start=2):
            mc = MORA_COLORS[dest_state]
            cell = ws.cell(row=cur_row, column=ci, value=dest_state)
            cell.fill = make_fill(COLOR_SUBHEAD)
            cell.font = make_font(bold=True, color=COLOR_WHITE, size=10)
            cell.alignment = make_align(horizontal="center")
            cell.border = thin_border()

        total_hdr = ws.cell(row=cur_row, column=6, value="TOTAL")
        total_hdr.fill = make_fill(COLOR_HEADER)
        total_hdr.font = make_font(bold=True, color=COLOR_WHITE, size=10)
        total_hdr.alignment = make_align(horizontal="center")
        total_hdr.border = thin_border()

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # Data rows
        for orig_state in all_states:
            mc_row = MORA_COLORS[orig_state]
            label_cell = ws.cell(row=cur_row, column=1, value=orig_state)
            label_cell.fill = make_fill(mc_row)
            label_cell.font = make_font(bold=True, size=10)
            label_cell.alignment = make_align(horizontal="left")
            label_cell.border = thin_border()

            row_total = int(pivot.loc[orig_state].sum())
            for ci, dest_state in enumerate(all_states, start=2):
                cnt = int(pivot.loc[orig_state, dest_state])
                mc_cell = MORA_COLORS[dest_state]
                pct_val = pct(cnt, row_total)
                display = f"{cnt:,} ({pct_val:.0f}%)" if row_total > 0 else "0"
                cell = ws.cell(row=cur_row, column=ci, value=display)
                cell.fill = make_fill(mc_cell)
                cell.font = make_font(size=9)
                cell.alignment = make_align(horizontal="right")
                cell.border = thin_border()

            # Row total
            rt_cell = ws.cell(row=cur_row, column=6, value=f"{row_total:,}")
            rt_cell.fill = make_fill(COLOR_TOTAL)
            rt_cell.font = make_font(bold=True, size=10)
            rt_cell.alignment = make_align(horizontal="right")
            rt_cell.border = thin_border()

            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        # TOTAL row
        total_label = ws.cell(row=cur_row, column=1, value="TOTAL")
        total_label.fill = make_fill(COLOR_TOTAL)
        total_label.font = make_font(bold=True, size=10)
        total_label.alignment = make_align(horizontal="right")
        total_label.border = thin_border()

        for ci, dest_state in enumerate(all_states, start=2):
            ct = col_totals[dest_state]
            pct_val = pct(ct, grand_total)
            display = f"{ct:,} ({pct_val:.0f}%)" if grand_total > 0 else "0"
            cell = ws.cell(row=cur_row, column=ci, value=display)
            cell.fill = make_fill(COLOR_TOTAL)
            cell.font = make_font(bold=True, size=9)
            cell.alignment = make_align(horizontal="right")
            cell.border = thin_border()

        gt_cell = ws.cell(row=cur_row, column=6, value=f"{grand_total:,}")
        gt_cell.fill = make_fill(COLOR_HEADER)
        gt_cell.font = make_font(bold=True, color=COLOR_WHITE, size=10)
        gt_cell.alignment = make_align(horizontal="right")
        gt_cell.border = thin_border()

        ws.row_dimensions[cur_row].height = 18
        cur_row += 3

    widths = {1:16, 2:20, 3:20, 4:20, 5:20, 6:16}
    set_col_widths(ws, widths)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Gráficas
# ══════════════════════════════════════════════════════════════════════════════

def write_graficas(wb: Workbook, data: dict):
    ws = wb.create_sheet("Gráficas")
    ws.sheet_view.showGridLines = False

    summary = data["summary"]
    n_camps = len(summary)

    # ── Data table: A1 title ───────────────────────────────────────────────
    style_header_cell(ws, 1, 1, "DATOS PARA GRÁFICAS", size=12, merge_to=13)
    ws.row_dimensions[1].height = 24

    headers = [
        "Campaña", "Total", "Nuevas", "De Anterior", "Persistentes", "Fugadas",
        "Inactiva", "Mora 1", "Mora 2", "Mora 3",
        "% Nuevas", "% Retención", "% Fuga",
    ]

    DATA_START_ROW = 3
    for ci, h in enumerate(headers, start=1):
        style_subheader_cell(ws, DATA_START_ROW, ci, h, size=10)
    ws.row_dimensions[DATA_START_ROW].height = 18

    prev_total = 0
    for ri, row in summary.iterrows():
        r = DATA_START_ROW + 1 + ri
        alt = ri % 2 == 0
        bg  = COLOR_ALT if alt else COLOR_WHITE
        pct_nuevas = pct(row["nuevas"], row["total"])
        pct_ret    = pct(row["de_anterior"], prev_total) if prev_total > 0 else 0.0
        pct_fuga   = pct(row["fugadas"], row["total"])
        prev_total = row["total"]

        vals = [
            row["camp_label"],
            row["total"],
            row["nuevas"],
            row["de_anterior"],
            row["persistentes"],
            row["fugadas"],
            row["Inactiva_total"],
            row["Mora 1_total"],
            row["Mora 2_total"],
            row["Mora 3_total"],
            pct_nuevas,
            pct_ret,
            pct_fuga,
        ]
        fmts = [None] + ["#,##0"] * 9 + ['0.0"%"'] * 3

        for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.fill = make_fill(bg)
            cell.font = make_font()
            cell.alignment = make_align(horizontal="center" if ci == 1 else "right")
            cell.border = thin_border()
            if fmt:
                cell.number_format = fmt

        ws.row_dimensions[r].height = 16

    DATA_END_ROW = DATA_START_ROW + n_camps  # inclusive

    # Row references for charts
    cats_ref = Reference(ws,
        min_col=1, max_col=1,
        min_row=DATA_START_ROW + 1, max_row=DATA_END_ROW)

    CHART_START_ROW = DATA_END_ROW + 3

    # ── Chart 1: Stacked bar — Mora composition ────────────────────────────
    chart1 = BarChart()
    chart1.type    = "col"
    chart1.grouping = "stacked"
    chart1.overlap = 100
    chart1.title   = "Composición por Estado de Mora (por Campaña)"
    chart1.y_axis.title = "Cuentas"
    chart1.x_axis.title = "Campaña"
    chart1.width   = 22
    chart1.height  = 14

    # Inactiva=col7, Mora1=col8, Mora2=col9, Mora3=col10
    for col_idx, (mora, color_hex) in enumerate(
        zip(MORA_LEVELS, [COLOR_INACTIVA, COLOR_MORA1, COLOR_MORA2, COLOR_MORA3]),
        start=7
    ):
        data_ref = Reference(ws,
            min_col=col_idx, max_col=col_idx,
            min_row=DATA_START_ROW, max_row=DATA_END_ROW)
        chart1.add_data(data_ref, titles_from_data=True)

    chart1.set_categories(cats_ref)
    ws.add_chart(chart1, f"A{CHART_START_ROW}")

    # ── Chart 2: Line — Nuevas vs De Anterior vs Fugadas ──────────────────
    chart2 = LineChart()
    chart2.title   = "Tendencias: Nuevas vs De Anterior vs Fugadas"
    chart2.y_axis.title = "Cuentas"
    chart2.x_axis.title = "Campaña"
    chart2.width   = 22
    chart2.height  = 14

    # Nuevas=col3, De Anterior=col4, Fugadas=col6
    for col_idx in [3, 4, 6]:
        data_ref = Reference(ws,
            min_col=col_idx, max_col=col_idx,
            min_row=DATA_START_ROW, max_row=DATA_END_ROW)
        chart2.add_data(data_ref, titles_from_data=True)

    chart2.set_categories(cats_ref)
    ws.add_chart(chart2, f"L{CHART_START_ROW}")

    # ── Chart 3: Stacked bar — Nuevas+De Anterior+Persistentes ────────────
    chart3 = BarChart()
    chart3.type     = "col"
    chart3.grouping = "stacked"
    chart3.overlap  = 100
    chart3.title    = "Composición: Nuevas + De Anterior + Persistentes"
    chart3.y_axis.title = "Cuentas"
    chart3.x_axis.title = "Campaña"
    chart3.width    = 22
    chart3.height   = 14

    # Nuevas=col3, De Anterior=col4, Persistentes=col5
    for col_idx in [3, 4, 5]:
        data_ref = Reference(ws,
            min_col=col_idx, max_col=col_idx,
            min_row=DATA_START_ROW, max_row=DATA_END_ROW)
        chart3.add_data(data_ref, titles_from_data=True)

    chart3.set_categories(cats_ref)

    CHART2_START_ROW = CHART_START_ROW + 24
    ws.add_chart(chart3, f"A{CHART2_START_ROW}")

    # ── Chart 4: Horizontal bar — % Fuga by campaign ──────────────────────
    chart4 = BarChart()
    chart4.type     = "bar"   # horizontal
    chart4.grouping = "clustered"
    chart4.title    = "% Fuga por Campaña (ventana 3 camps)"
    chart4.x_axis.title = "% Fuga"
    chart4.y_axis.title = "Campaña"
    chart4.width    = 22
    chart4.height   = 14

    fuga_ref = Reference(ws,
        min_col=13, max_col=13,
        min_row=DATA_START_ROW, max_row=DATA_END_ROW)
    chart4.add_data(fuga_ref, titles_from_data=True)
    chart4.set_categories(cats_ref)

    ws.add_chart(chart4, f"L{CHART2_START_ROW}")

    # Column widths
    widths = {i: 12 for i in range(1, 14)}
    widths[1] = 10
    set_col_widths(ws, widths)


# ══════════════════════════════════════════════════════════════════════════════
# CONSOLE OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def print_console_summary(data: dict):
    summary     = data["summary"]
    trans_df    = data["transitions"]
    pool_size   = len(data["pool"])

    print("\n" + "═" * 100)
    print("  TRACKING DE CARTERA — RESUMEN EJECUTIVO")
    print("═" * 100)

    # Summary table
    col_w = [6, 11, 10, 9, 13, 15, 12, 10, 10, 16]
    hdr = (
        f"{'Camp':^6} {'Pendientes':^11} {'% Pool':^10} {'Nuevas':^9} "
        f"{'De Anterior':^13} {'Persistentes':^15} {'Salen Sig':^12} "
        f"{'Fugadas':^10} {'% Fuga':^10} {'Saldo Total':^16}"
    )
    print(hdr)
    print("─" * 100)

    for _, row in summary.iterrows():
        pct_fuga = pct(row["fugadas"], row["total"])
        print(
            f"{row['camp_label']:^6} "
            f"{row['total']:>11,} "
            f"{row['pct_pool']:>9.1f}% "
            f"{row['nuevas']:>9,} "
            f"{row['de_anterior']:>13,} "
            f"{row['persistentes']:>15,} "
            f"{row['salen_siguiente']:>12,} "
            f"{row['fugadas']:>10,} "
            f"{pct_fuga:>9.1f}% "
            f"${row['saldo_total']:>15,.2f}"
        )

    print("═" * 100)

    # Transition snippet
    if not trans_df.empty:
        print("\n  MUESTRA DE TRANSICIONES (primeras 20 filas)")
        print("─" * 60)
        print(trans_df.head(20).to_string(index=False))

    # Key metrics
    print("\n" + "═" * 60)
    print("  MÉTRICAS CLAVE")
    print("─" * 60)
    print(f"  Pool total (NoDamas únicas):  {pool_size:,}")

    retenciones = []
    for i in range(1, len(summary)):
        prev_total = summary.iloc[i - 1]["total"]
        if prev_total > 0:
            retenciones.append(summary.iloc[i]["de_anterior"] / prev_total * 100)
    avg_ret = np.mean(retenciones) if retenciones else 0.0
    print(f"  Retención promedio entre camps: {avg_ret:.1f}%")

    summary["_pct_fuga"] = summary.apply(lambda r: pct(r["fugadas"], r["total"]), axis=1)
    max_fuga_idx  = summary["_pct_fuga"].idxmax()
    min_fuga_idx  = summary["_pct_fuga"].idxmin()
    print(f"  Campaña mayor % fuga: {summary.loc[max_fuga_idx,'camp_label']} "
          f"({summary.loc[max_fuga_idx,'_pct_fuga']:.1f}%)")
    print(f"  Campaña menor % fuga: {summary.loc[min_fuga_idx,'camp_label']} "
          f"({summary.loc[min_fuga_idx,'_pct_fuga']:.1f}%)")

    max_new_idx = summary["nuevas"].idxmax()
    print(f"  Campaña más nuevas: {summary.loc[max_new_idx,'camp_label']} "
          f"({summary.loc[max_new_idx,'nuevas']:,} nuevas)")
    print("═" * 60)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # Load & analyse
    df   = load_data()
    data = build_analysis(df)

    # Console output
    print_console_summary(data)

    # Build workbook
    print("\nGenerating Excel report …")
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_resumen_ejecutivo(wb, data)
    write_analisis_por_campana(wb, data)
    write_analisis_por_mora(wb, data)
    write_tracking_salidas(wb, data)
    write_matriz_transicion(wb, data)
    write_graficas(wb, data)

    # Save
    output_path = Path(OUTPUT_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    print(f"\n✓ Report saved: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print("Done.")


if __name__ == "__main__":
    main()
