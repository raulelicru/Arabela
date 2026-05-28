"""
analisis_inactivas_moras.py
────────────────────────────────────────────────────────────────────────────────
Análisis de flujo de cuentas:
    Pendientes Inactivas → Mora 1 → Mora 2 → Mora 3 → Salida

CATEGORÍAS POR CAMPAÑA:
  • Mora1 desde Inactivas  → en pendientes Y en Mora 1
  • Mora1 Nueva            → en Mora 1, SIN aparecer en pendientes
  • Nuevas M2-M3           → llegan directamente a Mora 2 o 3 (sin venir de M1 anterior)
  • Permanecen M2          → estaban en Mora 1 en campaña anterior, ahora en Mora 2
  • Permanecen M3          → estaban en Mora 2 en campaña anterior, ahora en Mora 3
  • No continúan           → estaban en campaña anterior, desaparecen en la actual

TOTAL CUADRA = activos_en_Cn + No_continúan
────────────────────────────────────────────────────────────────────────────────
INSTRUCCIONES DE USO:
  1. Ajusta las rutas en la sección CONFIGURACIÓN.
  2. Verifica los nombres de columnas.
  3. Ejecuta: python analisis_inactivas_moras.py
────────────────────────────────────────────────────────────────────────────────
"""

import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN  ←  MODIFICA AQUÍ
# ══════════════════════════════════════════════════════════════════════════════

PENDIENTES_FILE = Path("pendientes_inactivas.xlsx")
MORAS_FILE      = Path("base_moras.xlsx")
OUTPUT_FILE     = Path("reporte_flujo_moras.xlsx")

# ── Pendientes de Inactivas ────────────────────────────────────────────────
COL_NODAMA_PEND  = "NoDama"       # columna NoDama en el archivo de pendientes
# Si tu archivo tiene una columna de estatus y necesitas filtrar:
COL_STATUS_PEND  = None           # ej. "Estatus"  — None si no aplica filtro
STATUS_PENDIENTE = "Pendiente"    # valor que indica pendiente de pago

# ── Base de Moras ──────────────────────────────────────────────────────────
COL_NODAMA_MORA = "NoDama"
COL_CAMPANIA    = "Campania"      # valores esperados: C1, C2, ... C10
COL_MORA        = "Moras"         # valores esperados: Mora 1, Mora 2, Mora 3
COL_SALDO       = "SaldoDama"     # None si no existe o no quieres usarlo

MORA_SHEET      = None            # None = hoja con más filas; o ej. "Consolidado"

# ── Orden de campañas ──────────────────────────────────────────────────────
def _camp_order(v: str) -> int:
    try:
        return int(str(v).upper().replace("C", "").strip())
    except Exception:
        return 999

# ══════════════════════════════════════════════════════════════════════════════
#  COLORES Y ESTILOS
# ══════════════════════════════════════════════════════════════════════════════
C_HEADER   = "1A3C6E"
C_SUBHEAD  = "2563EB"
C_INAC     = "F3E8FF"   # morado pálido  → Mora1 desde Inactivas
C_M1NEW    = "DCFCE7"   # verde pálido   → Mora1 Nueva
C_M2NEW    = "DBEAFE"   # azul pálido    → Nuevas M2-M3
C_PERM2    = "FEF9C3"   # amarillo pálido → Permanecen M2
C_PERM3    = "FFEDD5"   # naranja pálido  → Permanecen M3
C_NOCONT   = "FEE2E2"   # rojo pálido    → No continúan
C_TOTAL    = "E0F2FE"   # azul claro     → Total
C_ALT      = "F8FAFC"
C_WHITE    = "FFFFFF"

def _fill(h):   return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _cell(ws, row, col, value, bold=False, fill=None, fc="000000",
          size=10, align="left", num_fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fc, size=size)
    c.alignment = _align(h=align, wrap=wrap)
    c.border    = _border()
    if fill:
        c.fill = _fill(fill)
    if num_fmt:
        c.number_format = num_fmt
    return c

def _hdr(ws, row, col, val, span=1, bg=C_HEADER, size=11):
    c = _cell(ws, row, col, val, bold=True, fill=bg, fc=C_WHITE,
              size=size, align="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return c

def _pct(a, b):
    return round(a / b * 100, 1) if b else 0.0

# ══════════════════════════════════════════════════════════════════════════════
#  CARGA DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

def _best_sheet(path: Path) -> str | None:
    xl = pd.ExcelFile(path)
    if len(xl.sheet_names) == 1:
        return xl.sheet_names[0]
    best, best_n = xl.sheet_names[0], 0
    for sh in xl.sheet_names:
        try:
            n = len(xl.parse(sh))
            if n > best_n:
                best, best_n = sh, n
        except Exception:
            pass
    return best


def load_pendientes(path: Path) -> set:
    """Carga el archivo de pendientes y devuelve un set de NoDamas."""
    print(f"  Cargando pendientes: {path.name}")
    sh = _best_sheet(path)
    df = pd.read_excel(path, sheet_name=sh)
    df.columns = [str(c).strip() for c in df.columns]

    # Buscar columna NoDama (flexible)
    nd_col = None
    for cand in [COL_NODAMA_PEND, "NoDama", "no_dama", "No Dama", "NODAMA"]:
        if cand in df.columns:
            nd_col = cand
            break
    if nd_col is None:
        # primer columna como fallback
        nd_col = df.columns[0]
        print(f"    AVISO: columna NoDama no encontrada, usando '{nd_col}'")

    df[nd_col] = df[nd_col].astype(str).str.strip()

    # Filtrar por status si aplica
    if COL_STATUS_PEND and COL_STATUS_PEND in df.columns:
        antes = len(df)
        df = df[df[COL_STATUS_PEND].astype(str).str.strip() == STATUS_PENDIENTE]
        print(f"    Filtro estatus '{STATUS_PENDIENTE}': {antes:,} → {len(df):,} registros")

    pendientes = set(df[nd_col].unique())
    print(f"    {len(pendientes):,} NoDamas pendientes de pago")
    return pendientes


def load_moras(path: Path) -> pd.DataFrame:
    """Carga la base de moras y la normaliza."""
    print(f"  Cargando moras: {path.name}")
    sh = MORA_SHEET or _best_sheet(path)
    df = pd.read_excel(path, sheet_name=sh)
    df.columns = [str(c).strip() for c in df.columns]

    # NoDama
    nd_col = None
    for cand in [COL_NODAMA_MORA, "NoDama", "no_dama", "No Dama"]:
        if cand in df.columns:
            nd_col = cand
            break
    if nd_col is None:
        raise ValueError("No se encontró columna NoDama en la base de moras.")

    # Campaña
    camp_col = None
    for cand in [COL_CAMPANIA, "Campania", "Campaña", "Camp", "campana"]:
        if cand in df.columns:
            camp_col = cand
            break
    if camp_col is None:
        raise ValueError("No se encontró columna Campaña en la base de moras.")

    # Mora
    mora_col = None
    for cand in [COL_MORA, "Moras", "Mora", "NivelMora", "nivel_mora"]:
        if cand in df.columns:
            mora_col = cand
            break
    if mora_col is None:
        raise ValueError("No se encontró columna Moras en la base de moras.")

    df[nd_col]   = df[nd_col].astype(str).str.strip()
    df[camp_col] = df[camp_col].astype(str).str.strip().str.upper()
    df[mora_col] = df[mora_col].astype(str).str.strip()

    # Normalizar mora
    mora_map = {
        "mora 1": "Mora 1", "mora1": "Mora 1", "1": "Mora 1",
        "mora 2": "Mora 2", "mora2": "Mora 2", "2": "Mora 2",
        "mora 3": "Mora 3", "mora3": "Mora 3", "3": "Mora 3",
    }
    df["_mora"] = df[mora_col].str.lower().map(mora_map)
    df = df.dropna(subset=["_mora"]).copy()

    if COL_SALDO and COL_SALDO in df.columns:
        df["_saldo"] = pd.to_numeric(df[COL_SALDO], errors="coerce").fillna(0)
    else:
        df["_saldo"] = 0

    df["_nodama"]  = df[nd_col]
    df["_camp"]    = df[camp_col]
    df["_camp_n"]  = df["_camp"].apply(_camp_order)
    df = df[df["_camp_n"] < 999].copy()

    print(f"    {len(df):,} registros | {df['_nodama'].nunique():,} NoDamas únicas | "
          f"{df['_camp'].nunique()} campañas")
    return df

# ══════════════════════════════════════════════════════════════════════════════
#  MOTOR DE ANÁLISIS
# ══════════════════════════════════════════════════════════════════════════════

def build_sets(df: pd.DataFrame) -> dict:
    """Construye sets de NoDamas por (campaña, mora)."""
    camps = sorted(df["_camp_n"].unique())
    sets = {}
    for c in camps:
        df_c = df[df["_camp_n"] == c]
        sets[c] = {
            "all":    set(df_c["_nodama"]),
            "Mora 1": set(df_c[df_c["_mora"] == "Mora 1"]["_nodama"]),
            "Mora 2": set(df_c[df_c["_mora"] == "Mora 2"]["_nodama"]),
            "Mora 3": set(df_c[df_c["_mora"] == "Mora 3"]["_nodama"]),
        }
    return sets, camps


def analyze_campaign(c, sets, camps, pendientes: set, i: int) -> dict:
    """Calcula todas las métricas para la campaña c."""
    cur   = sets[c]
    prev  = sets[camps[i - 1]] if i > 0 else None
    label = f"C{c}"

    m1 = cur["Mora 1"]
    m2 = cur["Mora 2"]
    m3 = cur["Mora 3"]
    all_cur = cur["all"]

    # ── Mora 1 ────────────────────────────────────────────────────────
    mora1_desde_inac = m1 & pendientes          # en M1 Y en pendientes
    mora1_nueva      = m1 - pendientes          # en M1 pero NO en pendientes

    # ── Mora 2 y 3 ───────────────────────────────────────────────────
    if prev:
        m1_prev = prev["Mora 1"]
        m2_prev = prev["Mora 2"]
        m3_prev = prev["Mora 3"]
        all_prev = prev["all"]

        permanecen_m2 = m2 & m1_prev            # M1 anterior → M2 actual
        permanecen_m3 = m3 & m2_prev            # M2 anterior → M3 actual
        nuevas_m2     = m2 - m1_prev            # llegan directo a M2
        nuevas_m3     = m3 - m2_prev            # llegan directo a M3
        no_continuan  = all_prev - all_cur      # desaparecen
    else:
        m1_prev = m2_prev = m3_prev = all_prev = set()
        permanecen_m2 = permanecen_m3 = set()
        nuevas_m2 = m2
        nuevas_m3 = m3
        no_continuan  = set()

    nuevas_m2_m3 = nuevas_m2 | nuevas_m3   # mutually exclusive (can't be in both)

    # Validación: activos = M1 + M2 + M3 (ya son mutually exclusive por definición)
    activos = len(all_cur)
    check   = (len(mora1_desde_inac) + len(mora1_nueva) +
               len(permanecen_m2)    + len(nuevas_m2)   +
               len(permanecen_m3)    + len(nuevas_m3))
    # activos debe == check (sets son disjuntos)

    total_reportado = activos + len(no_continuan)

    return {
        "camp_n":             c,
        "camp_label":         label,
        # Categorías principales
        "mora1_desde_inac":   len(mora1_desde_inac),
        "mora1_nueva":        len(mora1_nueva),
        "nuevas_m2_m3":       len(nuevas_m2_m3),
        "permanecen_m2":      len(permanecen_m2),
        "permanecen_m3":      len(permanecen_m3),
        "no_continuan":       len(no_continuan),
        # Desglose interno
        "nuevas_m2":          len(nuevas_m2),
        "nuevas_m3":          len(nuevas_m3),
        # Totales
        "total_mora1":        len(m1),
        "total_mora2":        len(m2),
        "total_mora3":        len(m3),
        "activos":            activos,
        "total_reportado":    total_reportado,
        "check_ok":           activos == check,
        # Saldos (si aplica)
        "_m1_ids":            m1,
        "_m2_ids":            m2,
        "_m3_ids":            m3,
        "_no_cont_ids":       no_continuan,
        "_mora1_inac_ids":    mora1_desde_inac,
        "_mora1_nueva_ids":   mora1_nueva,
    }


def run_analysis(pendientes: set, df_moras: pd.DataFrame) -> list[dict]:
    """Ejecuta el análisis completo para todas las campañas."""
    sets, camps = build_sets(df_moras)
    results = []
    for i, c in enumerate(camps):
        r = analyze_campaign(c, sets, camps, pendientes, i)
        results.append(r)
    return results


def build_summary(results: list[dict]) -> pd.DataFrame:
    cols = [
        "camp_label", "mora1_desde_inac", "mora1_nueva", "nuevas_m2_m3",
        "permanecen_m2", "permanecen_m3", "no_continuan", "total_reportado",
    ]
    rows = [{k: r[k] for k in cols} for r in results]
    df = pd.DataFrame(rows)
    df.columns = [
        "Campaña", "Mora1 desde Inactivas", "Mora1 Nuevas", "Nuevas M2-M3",
        "Permanecen M2", "Permanecen M3", "No continúan", "Total cuentas",
    ]
    return df


def build_permanence_metrics(results: list[dict]) -> pd.DataFrame:
    rows = []
    for i, r in enumerate(results):
        if i == 0:
            continue
        prev = results[i - 1]
        total_m1_prev = prev["total_mora1"]
        total_m2_prev = prev["total_mora2"]
        rows.append({
            "Transición":         f"{prev['camp_label']} → {r['camp_label']}",
            "M1 anterior":        total_m1_prev,
            "→ M2 (Permanecen)":  r["permanecen_m2"],
            "Tasa permanencia M2":f"{_pct(r['permanecen_m2'], total_m1_prev):.1f}%",
            "M2 anterior":        total_m2_prev,
            "→ M3 (Permanecen)":  r["permanecen_m3"],
            "Tasa permanencia M3":f"{_pct(r['permanecen_m3'], total_m2_prev):.1f}%",
            "No continúan":       r["no_continuan"],
            "Tasa fuga":          f"{_pct(r['no_continuan'], prev['activos']):.1f}%",
        })
    return pd.DataFrame(rows)


def build_fuga_metrics(results: list[dict]) -> pd.DataFrame:
    rows = []
    for i, r in enumerate(results):
        if i == 0:
            rows.append({
                "Campaña":         r["camp_label"],
                "Activos":         r["activos"],
                "No continúan":    0,
                "Tasa fuga":       "—",
                "Tasa permanencia":"—",
                "Tasa ingreso":    "—",
            })
            continue
        prev = results[i - 1]
        ingresos = r["mora1_desde_inac"] + r["mora1_nueva"] + r["nuevas_m2_m3"]
        rows.append({
            "Campaña":         r["camp_label"],
            "Activos":         r["activos"],
            "No continúan":    r["no_continuan"],
            "Tasa fuga":       f"{_pct(r['no_continuan'], prev['activos']):.1f}%",
            "Tasa permanencia":f"{_pct(r['permanecen_m2']+r['permanecen_m3'], prev['activos']):.1f}%",
            "Tasa ingreso":    f"{_pct(ingresos, r['activos']):.1f}%",
        })
    return pd.DataFrame(rows)

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — PESTAÑA 1: RESUMEN EJECUTIVO
# ══════════════════════════════════════════════════════════════════════════════

def write_resumen(wb: Workbook, results: list[dict], pendientes: set):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False

    col_widths = [12, 20, 14, 14, 16, 16, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"].value     = "ANÁLISIS DE FLUJO: INACTIVAS → MORA 1 → MORA 2 → MORA 3"
    ws["A1"].font      = Font(bold=True, size=14, color=C_WHITE)
    ws["A1"].fill      = _fill(C_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 32

    # Subtítulo
    ws.merge_cells("A2:H2")
    ws["A2"].value     = (f"Pool pendientes de Inactivas: {len(pendientes):,} cuentas  |  "
                          f"Campañas analizadas: C1–C{max(r['camp_n'] for r in results)}")
    ws["A2"].font      = Font(italic=True, size=10, color="64748B")
    ws["A2"].alignment = _align(h="center")

    row = 4

    # Leyenda de colores
    legend = [
        (C_INAC,   "Mora 1 desde Inactivas"),
        (C_M1NEW,  "Mora 1 Nueva"),
        (C_M2NEW,  "Nuevas M2-M3"),
        (C_PERM2,  "Permanecen M2"),
        (C_PERM3,  "Permanecen M3"),
        (C_NOCONT, "No continúan"),
    ]
    for ci, (color, label) in enumerate(legend):
        c = ws.cell(row=row, column=ci + 1, value=label)
        c.fill      = _fill(color)
        c.font      = Font(size=9, bold=True)
        c.alignment = _align(h="center")
        c.border    = _border()
    ws.row_dimensions[row].height = 16
    row += 2

    # Cabecera tabla
    headers = ["Campaña", "Mora1 desde\nInactivas", "Mora1\nNuevas",
               "Nuevas\nM2-M3", "Permanecen\nM2", "Permanecen\nM3",
               "No\ncontinúan", "Total\ncuentas"]
    col_fills = [C_HEADER, C_INAC, C_M1NEW, C_M2NEW, C_PERM2, C_PERM3, C_NOCONT, C_TOTAL]

    for ci, (h, bg) in enumerate(zip(headers, col_fills), 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(bold=True, color=C_WHITE if bg in (C_HEADER,) else "1A3C6E", size=10)
        c.fill      = _fill(bg)
        c.alignment = _align(h="center", wrap=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 30
    row += 1

    # Datos
    for ri, r in enumerate(results):
        bg_alt = C_ALT if ri % 2 == 0 else C_WHITE
        row_data = [
            r["camp_label"],
            r["mora1_desde_inac"],
            r["mora1_nueva"],
            r["nuevas_m2_m3"],
            r["permanecen_m2"],
            r["permanecen_m3"],
            r["no_continuan"],
            r["total_reportado"],
        ]
        row_colors = [bg_alt, C_INAC, C_M1NEW, C_M2NEW, C_PERM2, C_PERM3, C_NOCONT, C_TOTAL]
        for ci, (val, bg) in enumerate(zip(row_data, row_colors), 1):
            _cell(ws, row, ci, val, align="center",
                  fill=bg if bg not in (bg_alt,) else bg_alt,
                  bold=(ci == 8))
        row += 1

    # Totales
    row += 1
    _cell(ws, row, 1, "TOTAL", bold=True, fill=C_TOTAL, align="center")
    for ci in range(2, 9):
        key = ["mora1_desde_inac","mora1_nueva","nuevas_m2_m3",
               "permanecen_m2","permanecen_m3","no_continuan","total_reportado"][ci - 2]
        _cell(ws, row, ci, sum(r[key] for r in results),
              bold=True, fill=C_TOTAL, align="center")

    row += 3

    # Métricas clave
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).value     = "MÉTRICAS CLAVE"
    ws.cell(row=row, column=1).font      = Font(bold=True, size=11, color=C_WHITE)
    ws.cell(row=row, column=1).fill      = _fill(C_SUBHEAD)
    ws.cell(row=row, column=1).alignment = _align(h="center")
    row += 1

    total_activos    = sum(r["activos"]          for r in results)
    total_inac       = sum(r["mora1_desde_inac"] for r in results)
    total_m1new      = sum(r["mora1_nueva"]      for r in results)
    total_perm2      = sum(r["permanecen_m2"]    for r in results)
    total_perm3      = sum(r["permanecen_m3"]    for r in results)
    total_nocont     = sum(r["no_continuan"]     for r in results)
    total_m1         = sum(r["total_mora1"]      for r in results)

    kpis = [
        ("Pool total pendientes de Inactivas", f"{len(pendientes):,}"),
        ("Cuentas activas totales (todas las campañas)", f"{total_activos:,}"),
        ("Mora1 desde Inactivas (acumulado)", f"{total_inac:,} ({_pct(total_inac, total_m1):.1f}% de Mora1)"),
        ("Mora1 Nuevas (acumulado)", f"{total_m1new:,} ({_pct(total_m1new, total_m1):.1f}% de Mora1)"),
        ("Total Permanecen M2 (acumulado)", f"{total_perm2:,}"),
        ("Total Permanecen M3 (acumulado)", f"{total_perm3:,}"),
        ("Total No continúan (acumulado)", f"{total_nocont:,}"),
    ]
    for label, val in kpis:
        _cell(ws, row, 1, label, bold=True, fill=C_ALT, align="left")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        _cell(ws, row, 6, val, bold=False, fill=C_WHITE, align="left")
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        row += 1

    ws.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — PESTAÑA 2: DETALLE POR CAMPAÑA
# ══════════════════════════════════════════════════════════════════════════════

def write_detalle(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Detalle por Campaña")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 14

    ws.merge_cells("A1:G1")
    ws["A1"].value     = "DETALLE POR CAMPAÑA"
    ws["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
    ws["A1"].fill      = _fill(C_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    row = 3
    for r in results:
        tot    = r["activos"]
        total_r = r["total_reportado"]

        # Cabecera de campaña
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1).value     = f"  CAMPAÑA {r['camp_label']}"
        ws.cell(row=row, column=1).font      = Font(bold=True, size=12, color=C_WHITE)
        ws.cell(row=row, column=1).fill      = _fill(C_SUBHEAD)
        ws.cell(row=row, column=1).alignment = _align(h="left")
        ws.row_dimensions[row].height = 22
        row += 1

        # KPI fila
        kpi_data = [
            ("Total activos", tot),
            ("Total M1", r["total_mora1"]),
            ("Total M2", r["total_mora2"]),
            ("Total M3", r["total_mora3"]),
            ("No continúan", r["no_continuan"]),
            ("Total reportado", total_r),
        ]
        for ci, (label, val) in enumerate(kpi_data, 1):
            _cell(ws, row, ci, f"{label}: {val:,}", bold=(ci == 6),
                  fill=C_TOTAL if ci == 6 else C_ALT, align="center")
        row += 1

        # Tabla de categorías
        cat_headers = ["Categoría", "Cuentas", "% del Activo", "% del Total",
                       "M1", "M2", "M3"]
        for ci, h in enumerate(cat_headers, 1):
            _cell(ws, row, ci, h, bold=True, fill="374151", fc=C_WHITE, align="center")
        row += 1

        categories = [
            ("Mora1 desde Inactivas", r["mora1_desde_inac"], C_INAC,
             r["mora1_desde_inac"], 0, 0),
            ("Mora1 Nuevas",          r["mora1_nueva"],      C_M1NEW,
             r["mora1_nueva"],        0, 0),
            ("Nuevas M2",             r["nuevas_m2"],        C_M2NEW,
             0, r["nuevas_m2"],       0),
            ("Nuevas M3",             r["nuevas_m3"],        C_M2NEW,
             0, 0,                    r["nuevas_m3"]),
            ("Permanecen M2",         r["permanecen_m2"],    C_PERM2,
             0, r["permanecen_m2"],   0),
            ("Permanecen M3",         r["permanecen_m3"],    C_PERM3,
             0, 0,                    r["permanecen_m3"]),
            ("No continúan",          r["no_continuan"],     C_NOCONT,
             0, 0,                    0),
        ]
        for cat, cnt, bg, m1c, m2c, m3c in categories:
            _cell(ws, row, 1, cat,                              fill=bg)
            _cell(ws, row, 2, cnt,         align="center",     fill=bg)
            _cell(ws, row, 3, f"{_pct(cnt, tot):.1f}%",
                  align="center", fill=bg)
            _cell(ws, row, 4, f"{_pct(cnt, total_r):.1f}%",
                  align="center", fill=bg)
            _cell(ws, row, 5, m1c or "—", align="center", fill=bg)
            _cell(ws, row, 6, m2c or "—", align="center", fill=bg)
            _cell(ws, row, 7, m3c or "—", align="center", fill=bg)
            row += 1

        # Total fila
        _cell(ws, row, 1, "TOTAL ACTIVO", bold=True, fill=C_TOTAL)
        _cell(ws, row, 2, tot,            bold=True, fill=C_TOTAL, align="center")
        _cell(ws, row, 3, "100.0%",       bold=True, fill=C_TOTAL, align="center")
        _cell(ws, row, 4, f"{_pct(tot, total_r):.1f}%",
              bold=True, fill=C_TOTAL, align="center")
        _cell(ws, row, 5, r["total_mora1"], bold=True, fill=C_TOTAL, align="center")
        _cell(ws, row, 6, r["total_mora2"], bold=True, fill=C_TOTAL, align="center")
        _cell(ws, row, 7, r["total_mora3"], bold=True, fill=C_TOTAL, align="center")
        row += 3

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — PESTAÑA 3: VALIDACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def write_validacion(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Validación de Totales")
    ws.sheet_view.showGridLines = False
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 18

    ws.merge_cells("A1:H1")
    ws["A1"].value     = "VALIDACIÓN DE TOTALES — CUADRE POR CAMPAÑA"
    ws["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
    ws["A1"].fill      = _fill(C_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    row = 3
    hdrs = ["Campaña", "M1+M2+M3\n(activos)", "Suma de\nCategorías",
            "¿Cuadra?", "Total\nReportado", "No continúan",
            "Act + NoCont", "¿Total OK?"]
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, row, ci, h, bold=True, fill=C_HEADER, fc=C_WHITE,
              align="center", wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1

    for r in results:
        activos   = r["activos"]
        cat_sum   = (r["mora1_desde_inac"] + r["mora1_nueva"] + r["nuevas_m2_m3"] +
                     r["permanecen_m2"]    + r["permanecen_m3"])
        cuadra    = activos == cat_sum
        total_r   = r["total_reportado"]
        act_nc    = activos + r["no_continuan"]
        total_ok  = total_r == act_nc

        row_color = "DCFCE7" if (cuadra and total_ok) else "FEE2E2"
        vals = [r["camp_label"], activos, cat_sum,
                "✓ SÍ" if cuadra else "✗ ERROR",
                total_r, r["no_continuan"],
                act_nc, "✓ SÍ" if total_ok else "✗ ERROR"]
        for ci, v in enumerate(vals, 1):
            bold = ci in (4, 8)
            _cell(ws, row, ci, v, align="center",
                  fill=row_color if ci in (4, 8) else C_ALT, bold=bold)
        row += 1

    ws.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — PESTAÑA 4: PERMANENCIA Y FUGA
# ══════════════════════════════════════════════════════════════════════════════

def write_permanencia(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Permanencia y Fuga")
    ws.sheet_view.showGridLines = False
    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 18

    ws.merge_cells("A1:I1")
    ws["A1"].value     = "MÉTRICAS DE PERMANENCIA Y FUGA ENTRE CAMPAÑAS"
    ws["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
    ws["A1"].fill      = _fill(C_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    row = 3
    hdrs = ["Transición", "M1 anterior", "→ M2\n(Permanecen)",
            "Tasa\nPerm. M2", "M2 anterior", "→ M3\n(Permanecen)",
            "Tasa\nPerm. M3", "No continúan", "Tasa\nFuga"]
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, row, ci, h, bold=True, fill=C_HEADER, fc=C_WHITE,
              align="center", wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1

    for i, r in enumerate(results):
        if i == 0:
            continue
        prev = results[i - 1]
        m1_prev = prev["total_mora1"]
        m2_prev = prev["total_mora2"]
        vals = [
            f"{prev['camp_label']} → {r['camp_label']}",
            m1_prev,
            r["permanecen_m2"],
            f"{_pct(r['permanecen_m2'], m1_prev):.1f}%",
            m2_prev,
            r["permanecen_m3"],
            f"{_pct(r['permanecen_m3'], m2_prev):.1f}%",
            r["no_continuan"],
            f"{_pct(r['no_continuan'], prev['activos']):.1f}%",
        ]
        row_bg = C_ALT if (i % 2 == 0) else C_WHITE
        for ci, v in enumerate(vals, 1):
            bg = {3: C_PERM2, 4: C_PERM2, 5: C_PERM3, 6: C_PERM3,
                  7: C_NOCONT, 8: C_NOCONT}.get(ci, row_bg)
            _cell(ws, row, ci, v, align="center", fill=bg)
        row += 1

    # Promedio
    row += 1
    if len(results) > 1:
        prom_pm2 = np.mean([_pct(r["permanecen_m2"], results[i-1]["total_mora1"])
                            for i, r in enumerate(results) if i > 0])
        prom_pm3 = np.mean([_pct(r["permanecen_m3"], results[i-1]["total_mora2"])
                            for i, r in enumerate(results) if i > 0])
        prom_fuga= np.mean([_pct(r["no_continuan"],  results[i-1]["activos"])
                            for i, r in enumerate(results) if i > 0])
        _cell(ws, row, 1, "PROMEDIO", bold=True, fill=C_TOTAL)
        _cell(ws, row, 4, f"{prom_pm2:.1f}%", bold=True, fill=C_PERM2, align="center")
        _cell(ws, row, 7, f"{prom_pm3:.1f}%", bold=True, fill=C_PERM3, align="center")
        _cell(ws, row, 9, f"{prom_fuga:.1f}%",bold=True, fill=C_NOCONT, align="center")

    ws.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — PESTAÑA 5: GRÁFICAS
# ══════════════════════════════════════════════════════════════════════════════

def write_graficas(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Gráficas")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    ws["A1"].value     = "GRÁFICAS — FLUJO DE CUENTAS POR CAMPAÑA"
    ws["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
    ws["A1"].fill      = _fill(C_HEADER)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    # Tabla de datos
    dr = 3
    hdrs = ["Campaña", "Mora1 Inac.", "Mora1 Nueva", "Nuevas M2-M3",
            "Permanecen M2", "Permanecen M3", "No continúan", "Total",
            "Activos"]
    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=dr, column=ci, value=h).font = Font(bold=True, size=9)

    for ri, r in enumerate(results, 1):
        ws.cell(row=dr+ri, column=1).value = r["camp_label"]
        for ci, key in enumerate(
            ["mora1_desde_inac","mora1_nueva","nuevas_m2_m3",
             "permanecen_m2","permanecen_m3","no_continuan",
             "total_reportado","activos"], 2):
            ws.cell(row=dr+ri, column=ci).value = r[key]

    n  = len(results)
    de = dr + n

    # Gráfica 1: barras apiladas composición
    bar = BarChart()
    bar.type = "col"; bar.grouping = "stacked"
    bar.title = "Composición por Campaña"
    bar.y_axis.title = "Cuentas"; bar.x_axis.title = "Campaña"
    bar.width = 22; bar.height = 14
    cats = Reference(ws, min_col=1, min_row=dr+1, max_row=de)
    data = Reference(ws, min_col=2, max_col=7, min_row=dr, max_row=de)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar_clrs = ["7C3AED","16A34A","2563EB","EAB308","F97316","EF4444"]
    for i, clr in enumerate(bar_clrs):
        bar.series[i].graphicalProperties.solidFill = clr
    ws.add_chart(bar, "A3")

    # Gráfica 2: líneas permanencia
    line = LineChart()
    line.title = "Permanencia y Fuga"
    line.y_axis.title = "Cuentas"; line.x_axis.title = "Campaña"
    line.width = 22; line.height = 14; line.style = 10
    data2 = Reference(ws, min_col=5, max_col=7, min_row=dr, max_row=de)
    line.add_data(data2, titles_from_data=True)
    line.set_categories(cats)
    for i, clr in enumerate(["EAB308","F97316","EF4444"]):
        line.series[i].graphicalProperties.line.solidFill = clr
        line.series[i].graphicalProperties.line.width = 20000
    ws.add_chart(line, "A22")

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "═"*70)
    print("  ANÁLISIS DE FLUJO: INACTIVAS → MORA 1 → MORA 2 → MORA 3")
    print("═"*70)

    # ── Verificar archivos ────────────────────────────────────────────────
    missing = [f for f in [PENDIENTES_FILE, MORAS_FILE] if not f.exists()]
    if missing:
        print("\n⚠  Archivos de entrada no encontrados:")
        for f in missing:
            print(f"   • {f}")
        print("\n  → Modifica las rutas en la sección CONFIGURACIÓN del script.")
        print("  → Generando reporte de EJEMPLO con datos simulados...\n")
        _run_demo()
        return

    # ── Cargar datos ──────────────────────────────────────────────────────
    print("\n📂 Cargando datos...")
    pendientes = load_pendientes(PENDIENTES_FILE)
    df_moras   = load_moras(MORAS_FILE)

    # ── Análisis ──────────────────────────────────────────────────────────
    print("\n🔍 Calculando métricas por campaña...")
    results = run_analysis(pendientes, df_moras)

    # ── Consola ───────────────────────────────────────────────────────────
    _print_results(results, pendientes)

    # ── Excel ─────────────────────────────────────────────────────────────
    print("\n📊 Generando Excel...")
    _write_excel(results, pendientes)


def _print_results(results, pendientes):
    summary = build_summary(results)
    print("\n" + "═"*100)
    print(f"{'Campaña':<10} {'M1 Inac.':>10} {'M1 Nueva':>10} {'M2-M3 Nv.':>10} "
          f"{'Perm.M2':>10} {'Perm.M3':>10} {'No Cont.':>10} {'Total':>10}")
    print("─"*100)
    for _, row in summary.iterrows():
        print(f"{row['Campaña']:<10} "
              f"{row['Mora1 desde Inactivas']:>10,} "
              f"{row['Mora1 Nuevas']:>10,} "
              f"{row['Nuevas M2-M3']:>10,} "
              f"{row['Permanecen M2']:>10,} "
              f"{row['Permanecen M3']:>10,} "
              f"{row['No continúan']:>10,} "
              f"{row['Total cuentas']:>10,}")
    print("═"*100)

    total_row = summary.drop(columns=["Campaña"]).sum()
    print(f"{'TOTAL':<10} "
          f"{total_row['Mora1 desde Inactivas']:>10,.0f} "
          f"{total_row['Mora1 Nuevas']:>10,.0f} "
          f"{total_row['Nuevas M2-M3']:>10,.0f} "
          f"{total_row['Permanecen M2']:>10,.0f} "
          f"{total_row['Permanecen M3']:>10,.0f} "
          f"{total_row['No continúan']:>10,.0f} "
          f"{total_row['Total cuentas']:>10,.0f}")

    # Validación
    print("\n✅ VALIDACIÓN DE CUADRE:")
    all_ok = True
    for r in results:
        cat_sum = (r["mora1_desde_inac"] + r["mora1_nueva"] + r["nuevas_m2_m3"] +
                   r["permanecen_m2"] + r["permanecen_m3"])
        ok = r["activos"] == cat_sum
        if not ok:
            all_ok = False
            print(f"   ✗ {r['camp_label']}: activos={r['activos']:,} ≠ suma_categorías={cat_sum:,}")
    if all_ok:
        print("   ✓ Todos los totales cuadran perfectamente.")

    print(f"\n   Pool pendientes Inactivas : {len(pendientes):,}")
    print(f"   Total cuentas únicas Moras: {sum(r['activos'] for r in results):,}")


def _write_excel(results, pendientes):
    wb = Workbook()
    wb.remove(wb.active)
    write_resumen(wb, results, pendientes)
    write_detalle(wb, results)
    write_validacion(wb, results)
    write_permanencia(wb, results)
    write_graficas(wb, results)
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Archivo guardado: {OUTPUT_FILE}")
    print(f"   Pestañas: {[s.title for s in wb.worksheets]}")


# ══════════════════════════════════════════════════════════════════════════════
#  DEMO (cuando no hay archivos reales)
# ══════════════════════════════════════════════════════════════════════════════

def _run_demo():
    """Genera datos simulados y produce el reporte de ejemplo."""
    import random
    random.seed(42)

    # Simular pendientes de inactivas
    pendientes = {f"D{i:05d}" for i in range(1, 1001)}

    # Simular base de moras (10 campañas)
    moras_rows = []
    pool = list(pendientes) + [f"N{i:05d}" for i in range(1, 2001)]

    # Model proper account progression: each account is in at most ONE mora per campaign
    # State tracking: "free" | "m1" | "m2" | "m3" | "exited"
    account_state = {}   # NoDama → current state
    pend_list = list(pendientes)
    new_pool  = [x for x in pool if x not in pendientes]

    prev_m1: set = set()
    prev_m2: set = set()

    for camp_n in range(1, 11):
        camp_label = f"C{camp_n}"

        # 1. Progress prev_m2 → m3 first (60%)
        m3_from_m2 = set(random.sample(list(prev_m2), int(len(prev_m2) * 0.60))) if prev_m2 else set()
        m3_new     = set(random.sample(new_pool, random.randint(2, 4)))
        m3 = m3_from_m2 | m3_new

        # 2. Progress prev_m1 → m2 (accounts NOT in m3; 55%)
        available_m1 = [x for x in prev_m1 if x not in m3]
        m2_from_m1  = set(random.sample(available_m1, int(len(available_m1) * 0.55))) if available_m1 else set()
        m2_new      = set(random.sample([x for x in new_pool if x not in m3], random.randint(3, 5)))
        m2 = m2_from_m1 | m2_new

        # 3. New M1: fresh pendientes + new accounts (none of which are in m2 or m3)
        occupied = m2 | m3
        pend_avail = [x for x in pend_list if x not in occupied]
        m1_inac    = set(random.sample(pend_avail, min(65, len(pend_avail))))
        new_avail  = [x for x in new_pool if x not in occupied and x not in m1_inac]
        m1_new     = set(random.sample(new_avail, min(random.randint(10, 20), len(new_avail))))
        m1 = m1_inac | m1_new

        # Verify disjoint (debug)
        assert not (m1 & m2), f"C{camp_n}: m1∩m2 not empty"
        assert not (m1 & m3), f"C{camp_n}: m1∩m3 not empty"
        assert not (m2 & m3), f"C{camp_n}: m2∩m3 not empty"

        for nd in m1:
            moras_rows.append({"NoDama": nd, "Campania": camp_label,
                                "Moras": "Mora 1", "SaldoDama": random.uniform(100, 5000)})
        for nd in m2:
            moras_rows.append({"NoDama": nd, "Campania": camp_label,
                                "Moras": "Mora 2", "SaldoDama": random.uniform(100, 5000)})
        for nd in m3:
            moras_rows.append({"NoDama": nd, "Campania": camp_label,
                                "Moras": "Mora 3", "SaldoDama": random.uniform(100, 5000)})

        prev_m1, prev_m2 = m1, m2

    df_moras = pd.DataFrame(moras_rows)
    df_moras["_mora"]   = df_moras["Moras"]
    df_moras["_nodama"] = df_moras["NoDama"]
    df_moras["_camp"]   = df_moras["Campania"]
    df_moras["_camp_n"] = df_moras["_camp"].apply(_camp_order)
    df_moras["_saldo"]  = df_moras["SaldoDama"]

    results = run_analysis(pendientes, df_moras)
    print("\n📋 Reporte de ejemplo (datos simulados):")
    _print_results(results, pendientes)

    demo_output = Path("reporte_flujo_moras_DEMO.xlsx")
    global OUTPUT_FILE
    orig = OUTPUT_FILE
    OUTPUT_FILE = demo_output
    _write_excel(results, pendientes)
    OUTPUT_FILE = orig


# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    main()
